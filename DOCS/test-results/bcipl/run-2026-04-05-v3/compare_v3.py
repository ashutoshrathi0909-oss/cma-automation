"""Compare v3 generated CMA Excel against ground truth. READ-ONLY."""
import openpyxl
import xlrd
import json
import os

# ── File paths ──
BASE = os.path.dirname(os.path.abspath(__file__))
gen_path = os.path.join(BASE, "BCIPL_generated_CMA_newpipeline.xlsm")
gt_path = "C:/Users/ASHUTOSH/OneDrive/Desktop/FInancial Analysis/BCIPL/CMA BCIPL 12012024.xls"

assert os.path.exists(gen_path), f"Generated not found: {gen_path}"
assert os.path.exists(gt_path), f"GT not found: {gt_path}"
print(f"Generated: {os.path.getsize(gen_path)} bytes")
print(f"Ground truth: {os.path.getsize(gt_path)} bytes")

# ── Open workbooks ──
wb_gen_f = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=False)
wb_gen_v = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=True)
wb_gt = xlrd.open_workbook(gt_path)

ws_gen_f = wb_gen_f["INPUT SHEET"]
ws_gen_v = wb_gen_v["INPUT SHEET"]
ws_gt = wb_gt.sheet_by_name("INPUT SHEET")

# ── Structure checks ──
print("\n=== STRUCTURE VERIFICATION ===")
print(f"VBA macros: {wb_gen_f.vba_archive is not None}")
print(f"Sheets: {wb_gen_f.sheetnames}")
print(f"Client (A7): {ws_gen_v.cell(row=7, column=1).value}")
years = [ws_gen_v.cell(row=8, column=c).value for c in range(2, 5)]
print(f"Years (B8-D8): {years}")
nature = [ws_gen_v.cell(row=10, column=c).value for c in range(2, 5)]
print(f"Nature (B10-D10): {nature}")
units = ws_gen_v.cell(row=13, column=2).value
print(f"Units (B13): {repr(units)}")
proj_years = [ws_gen_v.cell(row=8, column=c).value for c in range(5, 9)]
print(f"Projection years (E8-H8): {proj_years}")

# ── Data-entry rows from cma_field_rows.py ──
PNL = {
    "Domestic Sales": 22, "Export Sales": 23, "Less Excise Duty and Cess": 25,
    "Dividends received from Mutual Funds": 29, "Interest Received": 30,
    "Profit on sale of fixed assets / Investments": 31, "Gain on Exchange Fluctuations": 32,
    "Extraordinary income": 33, "Others (Non-Operating Income)": 34,
    "Raw Materials Consumed (Imported)": 41, "Raw Materials Consumed (Indigenous)": 42,
    "Stores and spares consumed (Imported)": 43, "Stores and spares consumed (Indigenous)": 44,
    "Wages": 45, "Processing / Job Work Charges": 46, "Freight and Transportation Charges": 47,
    "Power, Coal, Fuel and Water": 48, "Others (Manufacturing)": 49,
    "Repairs & Maintenance (Manufacturing)": 50, "Security Service Charges": 51,
    "Stock in process Opening Balance": 53, "Stock in process Closing Balance": 54,
    "Depreciation (Manufacturing)": 56, "Finished Goods Opening Balance": 58,
    "Finished Goods Closing Balance": 59, "Depreciation (CMA)": 63,
    "Other Manufacturing Exp (CMA)": 64, "Salary and staff expenses": 67,
    "Rent, Rates and Taxes": 68, "Bad Debts": 69,
    "Advertisements and Sales Promotions": 70, "Others (Admin)": 71,
    "Repairs & Maintenance (Admin)": 72, "Audit Fees & Directors Remuneration": 73,
    "Miscellaneous Expenses written off": 75, "Deferred Revenue Expenditures": 76,
    "Other Amortisations": 77, "Interest on Fixed Loans / Term loans": 83,
    "Interest on Working capital loans": 84, "Bank Charges": 85,
    "Loss on sale of fixed assets / Investments": 89, "Sundry Balances Written off": 90,
    "Loss on Exchange Fluctuations": 91, "Extraordinary losses": 92,
    "Others (Non-Operating Expenses)": 93, "Income Tax provision": 99,
    "Deferred Tax Liability (P&L)": 100, "Deferred Tax Asset (P&L)": 101,
    "Brought forward from previous year": 106, "Dividend": 107,
    "Other Appropriation of profit": 108,
}
BS = {
    "Issued, Subscribed and Paid up": 116, "Share Application Money": 117,
    "General Reserve": 121, "Balance transferred from profit and loss a/c": 122,
    "Share Premium A/c": 123, "Revaluation Reserve": 124, "Other Reserve": 125,
    "Working Capital Bank Finance - Bank 1": 131, "Working Capital Bank Finance - Bank 2": 132,
    "Term Loan Repayable in next one year": 136, "Term Loan Balance Repayable after one year": 137,
    "Debentures Repayable in next one year": 140, "Debentures Balance Repayable after one year": 141,
    "Preference Shares Repayable in next one year": 144, "Preference Shares Balance Repayable after one year": 145,
    "Other Debts Repayable in Next One year": 148, "Balance Other Debts": 149,
    "Unsecured Loans - Quasi Equity": 152, "Unsecured Loans - Long Term Debt": 153,
    "Unsecured Loans - Short Term Debt": 154, "Deferred tax liability (BS)": 159,
    "Gross Block": 162, "Less Accumulated Depreciation": 163, "Capital Work in Progress": 165,
    "Patents / goodwill / copyrights etc": 169, "Misc Expenditure (to the extent not w/o)": 170,
    "Deferred Tax Asset (BS)": 171, "Other Intangible assets": 172,
    "Additions to Fixed Assets": 175, "Sale of Fixed assets WDV": 176,
    "Profit on sale of Fixed assets (BS)": 177, "Loss on sale of Fixed assets (BS)": 178,
    "Investment in Govt. Securities (Current)": 182, "Investment in Govt. Securities (Non Current)": 183,
    "Other current investments": 185, "Other non current investments": 186,
    "Investment in group companies / subsidiaries": 188,
    "Raw Material Imported": 193, "Raw Material Indigenous": 194,
    "Stores and Spares Imported": 197, "Stores and Spares Indigenous": 198,
    "Stocks-in-process": 200, "Finished Goods": 201,
    "Domestic Receivables": 206, "Export Receivables": 207, "Debtors more than 6 months": 208,
    "Cash on Hand": 212, "Bank Balances": 213, "Fixed Deposit under lien": 214,
    "Other Fixed Deposits": 215, "Advances recoverable in cash or in kind": 219,
    "Advances to suppliers of raw materials": 220, "Advance Income Tax": 221,
    "Prepaid Expenses": 222, "Other Advances / current asset": 223,
    "Advances to group / subsidiaries companies": 224,
    "Exposure in group companies - Investments": 229, "Exposure in group companies - Advances": 230,
    "Debtors more than six months (Non-Current)": 232, "Investments (Non-Current)": 233,
    "Fixed Deposits (Non Current)": 234, "Dues from directors / partners / promoters": 235,
    "Advances to suppliers of capital goods": 236,
    "Security deposits with government departments": 237, "Other non current assets": 238,
    "Sundry Creditors for goods": 242, "Advance received from customers": 243,
    "Provision for Taxation": 244, "Dividend payable": 245,
    "Other statutory liabilities (due within 1 year)": 246,
    "Interest Accrued but not due": 247, "Interest Accrued and due": 248,
    "Creditors for Expenses": 249, "Other current liabilities": 250,
    "Arrears of cumulative dividends": 254, "Gratuity liability not provided for": 255,
    "Disputed excise / customs / tax liabilities": 256,
    "Bank guarantee / Letter of credit outstanding": 257, "Other contingent liabilities": 258,
}
ALL_FIELDS = {**PNL, **BS}
ROW_TO_NAME = {v: k for k, v in ALL_FIELDS.items()}


def compare_cell(expected, actual, tol=0.02):
    e_empty = (expected is None or expected == ""
               or (isinstance(expected, (int, float)) and abs(expected) < 1e-9))
    a_empty = (actual is None or actual == ""
               or (isinstance(actual, (int, float)) and abs(actual) < 1e-9))
    if e_empty and a_empty:
        return ("BOTH_EMPTY", 0)
    if e_empty and not a_empty:
        return ("EXTRA", None)
    if not e_empty and a_empty:
        return ("MISSING", None)
    if not isinstance(expected, (int, float)) or not isinstance(actual, (int, float)):
        return ("NON_NUMERIC", None)
    if abs(expected) < 0.001:
        if abs(actual) < 0.001:
            return ("MATCH", 0)
        return ("MISMATCH", float("inf"))
    diff = abs(actual - expected) / abs(expected)
    if diff <= tol:
        return ("MATCH", diff)
    return ("MISMATCH", diff)


# ── Read all data-entry cells ──
COL_MAP = {2: "B(2021)", 3: "C(2022)", 4: "D(2023)"}
results = []
for fname, row in sorted(ALL_FIELDS.items(), key=lambda x: x[1]):
    for col in [2, 3, 4]:
        gt_val = ws_gt.cell_value(row - 1, col - 1)
        gen_val = ws_gen_v.cell(row=row, column=col).value
        gen_formula = ws_gen_f.cell(row=row, column=col).value

        if isinstance(gt_val, str) and gt_val.strip() == "":
            gt_val = None
        if isinstance(gen_val, str) and gen_val.strip() == "":
            gen_val = None

        status, diff = compare_cell(gt_val, gen_val)

        formula_str = ""
        if gen_formula is not None and isinstance(gen_formula, str) and gen_formula.startswith("="):
            formula_str = gen_formula
        elif gen_formula is not None and gen_formula != gen_val:
            formula_str = str(gen_formula)

        results.append({
            "row": row, "field": fname, "fy": COL_MAP[col],
            "expected": gt_val, "got": gen_val, "formula": formula_str,
            "status": status, "diff": diff,
        })

# ── Summarize ──
gt_nonzero = [r for r in results if r["status"] in ("MATCH", "MISMATCH", "MISSING")]
matches = [r for r in results if r["status"] == "MATCH"]
mismatches = [r for r in results if r["status"] == "MISMATCH"]
missing = [r for r in results if r["status"] == "MISSING"]
extra = [r for r in results if r["status"] == "EXTRA"]
both_empty = [r for r in results if r["status"] == "BOTH_EMPTY"]

print(f"\n=== DATA-ENTRY ACCURACY ===")
print(f"Total cells checked: {len(results)}")
print(f"GT has value: {len(gt_nonzero)}")
print(f"MATCH (within 2%): {len(matches)}")
print(f"MISMATCH: {len(mismatches)}")
print(f"MISSING: {len(missing)}")
print(f"EXTRA: {len(extra)}")
print(f"BOTH_EMPTY: {len(both_empty)}")
acc = len(matches) / len(gt_nonzero) * 100 if gt_nonzero else 0
print(f"ACCURACY: {acc:.1f}% ({len(matches)}/{len(gt_nonzero)})")

print(f"\n=== MATCHING CELLS ({len(matches)}) ===")
for r in sorted(matches, key=lambda x: (x["row"], x["fy"])):
    exp = f"{r['expected']:.4f}" if isinstance(r["expected"], float) else str(r["expected"])
    got = f"{r['got']:.4f}" if isinstance(r["got"], float) else str(r["got"])
    d = f"{r['diff']*100:.1f}%" if r["diff"] is not None else ""
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | exp={exp:>10s} got={got:>10s} | {d}")

print(f"\n=== MISMATCHED CELLS ({len(mismatches)}) ===")
for r in sorted(mismatches, key=lambda x: (x["row"], x["fy"])):
    exp = f"{r['expected']:.4f}" if isinstance(r["expected"], float) else str(r["expected"])
    got_raw = r["got"]
    got = f"{got_raw:.4f}" if isinstance(got_raw, float) else str(got_raw)
    d = f"{r['diff']*100:.0f}%" if r["diff"] is not None and r["diff"] != float("inf") else "INF"
    f = r["formula"][:80] if r["formula"] else ""
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | exp={exp:>10s} got={got:>10s} | diff={d:>8s} | {f}")

print(f"\n=== MISSING CELLS ({len(missing)}) ===")
for r in sorted(missing, key=lambda x: (x["row"], x["fy"])):
    exp = f"{r['expected']:.4f}" if isinstance(r["expected"], float) else str(r["expected"])
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | exp={exp}")

print(f"\n=== EXTRA CELLS ({len(extra)}) ===")
for r in sorted(extra, key=lambda x: (x["row"], x["fy"])):
    got = f"{r['got']:.4f}" if isinstance(r["got"], float) else str(r["got"])
    f = r["formula"][:80] if r["formula"] else ""
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | got={got} | {f}")

# ── Unit conversion check ──
print(f"\n=== UNIT CONVERSION CHECK (10 key rows) ===")
key_rows = [22, 42, 45, 83, 99, 116, 162, 206, 213, 242]
for row in key_rows:
    name = ROW_TO_NAME.get(row, "?")
    gt_b = ws_gt.cell_value(row - 1, 1)
    gt_c = ws_gt.cell_value(row - 1, 2)
    gt_d = ws_gt.cell_value(row - 1, 3)
    gen_b = ws_gen_v.cell(row=row, column=2).value
    gen_c = ws_gen_v.cell(row=row, column=3).value
    gen_d = ws_gen_v.cell(row=row, column=4).value

    def fmt(v):
        if v is None or v == "":
            return "EMPTY"
        if isinstance(v, float):
            return f"{v:.2f}"
        return str(v)

    print(f"  Row {row:3d} {name[:30]:30s} | GT=[{fmt(gt_b):>10s}, {fmt(gt_c):>10s}, {fmt(gt_d):>10s}] | Gen=[{fmt(gen_b):>10s}, {fmt(gen_c):>10s}, {fmt(gen_d):>10s}]")

    # Show formulas for populated cells
    for ci, col_letter in [(2, "B"), (3, "C"), (4, "D")]:
        gf = ws_gen_f.cell(row=row, column=ci).value
        if gf and isinstance(gf, str) and gf.startswith("="):
            print(f"    {col_letter} formula: {gf[:120]}")

# ── Save raw JSON ──
output = {
    "total": len(results),
    "gt_nonzero": len(gt_nonzero),
    "matches": len(matches),
    "mismatches": len(mismatches),
    "missing": len(missing),
    "extra": len(extra),
    "both_empty": len(both_empty),
    "accuracy": acc,
    "match_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"], "got": r["got"],
         "diff": r["diff"]}
        for r in matches
    ],
    "mismatch_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"], "got": r["got"],
         "diff": r["diff"] if r["diff"] != float("inf") else 99999,
         "formula": r["formula"]}
        for r in mismatches
    ],
    "missing_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"]}
        for r in missing
    ],
    "extra_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "got": r["got"], "formula": r["formula"]}
        for r in extra
    ],
}
out_path = os.path.join(BASE, "comparison-raw.json")
with open(out_path, "w") as f:
    json.dump(output, f, indent=2, default=str)
print(f"\nRaw results saved to {out_path}")
