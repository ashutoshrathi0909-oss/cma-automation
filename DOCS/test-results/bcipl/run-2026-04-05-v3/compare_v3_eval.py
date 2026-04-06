"""Compare v3 generated CMA Excel against ground truth.
Evaluates simple formulas (=A+B+C) since the file has no cached values.
READ-ONLY — does not modify any source files."""
import openpyxl
import xlrd
import json
import os
import re

BASE = os.path.dirname(os.path.abspath(__file__))
gen_path = os.path.join(BASE, "BCIPL_generated_CMA_newpipeline.xlsm")
gt_path = "C:/Users/ASHUTOSH/OneDrive/Desktop/FInancial Analysis/BCIPL/CMA BCIPL 12012024.xls"

wb_gen_f = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=False)
wb_gen_v = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=True)
wb_gt = xlrd.open_workbook(gt_path)

ws_gen_f = wb_gen_f["INPUT SHEET"]
ws_gen_v = wb_gen_v["INPUT SHEET"]
ws_gt = wb_gt.sheet_by_name("INPUT SHEET")

# ── Structure checks ──
print("=== STRUCTURE VERIFICATION ===")
vba_ok = wb_gen_f.vba_archive is not None
client = ws_gen_v.cell(row=7, column=1).value
years = [ws_gen_v.cell(row=8, column=c).value for c in range(2, 5)]
nature = [ws_gen_v.cell(row=10, column=c).value for c in range(2, 5)]
units = ws_gen_v.cell(row=13, column=2).value
proj_years = [ws_gen_v.cell(row=8, column=c).value for c in range(5, 9)]

struct_checks = [
    ("File format (.xlsm)", "PASS", "Correct format"),
    ("VBA macros present", "PASS" if vba_ok else "FAIL", str(vba_ok)),
    ("INPUT SHEET exists", "PASS", "Present"),
    ("Client name (Row 7)", "PASS" if client and "BCIPL" in str(client).upper() else "FAIL", str(client)),
    ("Years (Row 8 B-D)", "PASS" if years == [2021, 2022, 2023] else "FAIL", str(years)),
    ("Nature (Row 10 B-D)", "PASS" if all(str(n).lower() == "audited" for n in nature if n) else "FAIL", str(nature)),
    ("Units (Row 13 B)", "CHECK" if units and "crs" not in str(units).lower() else "PASS", repr(units)),
    ("Projection years (E-H)", "PASS" if proj_years == [2024, 2025, 2026, 2027] else "FAIL", str(proj_years)),
]
for name, result, detail in struct_checks:
    print(f"  {result:5s} | {name:35s} | {detail}")


def eval_formula(formula_str):
    """Evaluate simple formulas like =A+B-C or =SUM(X:Y).
    Returns computed float or None if can't evaluate."""
    if formula_str is None:
        return None
    if not isinstance(formula_str, str):
        return formula_str  # Already a number
    if not formula_str.startswith("="):
        try:
            return float(formula_str)
        except (ValueError, TypeError):
            return None

    expr = formula_str[1:]  # Strip =

    # Handle =0
    if expr.strip() == "0":
        return 0.0

    # Handle SUM references like =SUM(C75:C77) — can't evaluate without context
    if "SUM(" in expr.upper() or ":" in expr:
        return None

    # Handle cell references like =C214
    if re.match(r"^[A-Z]+\d+$", expr.strip()):
        return None  # Can't resolve cell refs

    # Handle simple arithmetic: =1.5+2.3-0.5
    # First check if it only contains numbers, +, -, *, and whitespace
    if re.match(r"^[\d.+\-*/ ]+$", expr):
        try:
            return float(eval(expr))  # Safe: only numbers and operators
        except Exception:
            return None

    # Handle expressions with cell refs mixed in — skip
    if re.search(r"[A-Z]+\d+", expr):
        return None

    try:
        return float(eval(expr))
    except Exception:
        return None


# ── Data-entry rows ──
PNL = {
    "Domestic Sales": 22, "Export Sales": 23, "Less Excise Duty and Cess": 25,
    "Dividends received from Mutual Funds": 29, "Interest Received": 30,
    "Profit on sale of fixed assets / Investments": 31, "Gain on Exchange Fluctuations": 32,
    "Extraordinary income": 33, "Others (Non-Operating Income)": 34,
    "Raw Materials Consumed (Imported)": 41, "Raw Materials Consumed (Indigenous)": 42,
    "Stores and spares consumed (Imported)": 43, "Stores and spares consumed (Indigenous)": 44,
    "Wages": 45, "Processing / Job Work Charges": 46, "Freight and Transportation Charges": 47,
    "Power, Coal, Fuel and Water": 48, "Others (Manufacturing)": 49,
    "Repairs & Maintenance (Manufacturing)": 50, "Security Service Charges": 51,
    "Stock in process Opening Balance": 53, "Stock in process Closing Balance": 54,
    "Depreciation (Manufacturing)": 56, "Finished Goods Opening Balance": 58,
    "Finished Goods Closing Balance": 59, "Depreciation (CMA)": 63,
    "Other Manufacturing Exp (CMA)": 64, "Salary and staff expenses": 67,
    "Rent, Rates and Taxes": 68, "Bad Debts": 69,
    "Advertisements and Sales Promotions": 70, "Others (Admin)": 71,
    "Repairs & Maintenance (Admin)": 72, "Audit Fees & Directors Remuneration": 73,
    "Miscellaneous Expenses written off": 75, "Deferred Revenue Expenditures": 76,
    "Other Amortisations": 77, "Interest on Fixed Loans / Term loans": 83,
    "Interest on Working capital loans": 84, "Bank Charges": 85,
    "Loss on sale of fixed assets / Investments": 89, "Sundry Balances Written off": 90,
    "Loss on Exchange Fluctuations": 91, "Extraordinary losses": 92,
    "Others (Non-Operating Expenses)": 93, "Income Tax provision": 99,
    "Deferred Tax Liability (P&L)": 100, "Deferred Tax Asset (P&L)": 101,
    "Brought forward from previous year": 106, "Dividend": 107,
    "Other Appropriation of profit": 108,
}
BS = {
    "Issued, Subscribed and Paid up": 116, "Share Application Money": 117,
    "General Reserve": 121, "Balance transferred from profit and loss a/c": 122,
    "Share Premium A/c": 123, "Revaluation Reserve": 124, "Other Reserve": 125,
    "Working Capital Bank Finance - Bank 1": 131, "Working Capital Bank Finance - Bank 2": 132,
    "Term Loan Repayable in next one year": 136, "Term Loan Balance Repayable after one year": 137,
    "Debentures Repayable in next one year": 140, "Debentures Balance Repayable after one year": 141,
    "Preference Shares Repayable in next one year": 144, "Preference Shares Balance Repayable after one year": 145,
    "Other Debts Repayable in Next One year": 148, "Balance Other Debts": 149,
    "Unsecured Loans - Quasi Equity": 152, "Unsecured Loans - Long Term Debt": 153,
    "Unsecured Loans - Short Term Debt": 154, "Deferred tax liability (BS)": 159,
    "Gross Block": 162, "Less Accumulated Depreciation": 163, "Capital Work in Progress": 165,
    "Patents / goodwill / copyrights etc": 169, "Misc Expenditure (to the extent not w/o)": 170,
    "Deferred Tax Asset (BS)": 171, "Other Intangible assets": 172,
    "Additions to Fixed Assets": 175, "Sale of Fixed assets WDV": 176,
    "Profit on sale of Fixed assets (BS)": 177, "Loss on sale of Fixed assets (BS)": 178,
    "Investment in Govt. Securities (Current)": 182, "Investment in Govt. Securities (Non Current)": 183,
    "Other current investments": 185, "Other non current investments": 186,
    "Investment in group companies / subsidiaries": 188,
    "Raw Material Imported": 193, "Raw Material Indigenous": 194,
    "Stores and Spares Imported": 197, "Stores and Spares Indigenous": 198,
    "Stocks-in-process": 200, "Finished Goods": 201,
    "Domestic Receivables": 206, "Export Receivables": 207, "Debtors more than 6 months": 208,
    "Cash on Hand": 212, "Bank Balances": 213, "Fixed Deposit under lien": 214,
    "Other Fixed Deposits": 215, "Advances recoverable in cash or in kind": 219,
    "Advances to suppliers of raw materials": 220, "Advance Income Tax": 221,
    "Prepaid Expenses": 222, "Other Advances / current asset": 223,
    "Advances to group / subsidiaries companies": 224,
    "Exposure in group companies - Investments": 229, "Exposure in group companies - Advances": 230,
    "Debtors more than six months (Non-Current)": 232, "Investments (Non-Current)": 233,
    "Fixed Deposits (Non Current)": 234, "Dues from directors / partners / promoters": 235,
    "Advances to suppliers of capital goods": 236,
    "Security deposits with government departments": 237, "Other non current assets": 238,
    "Sundry Creditors for goods": 242, "Advance received from customers": 243,
    "Provision for Taxation": 244, "Dividend payable": 245,
    "Other statutory liabilities (due within 1 year)": 246,
    "Interest Accrued but not due": 247, "Interest Accrued and due": 248,
    "Creditors for Expenses": 249, "Other current liabilities": 250,
    "Arrears of cumulative dividends": 254, "Gratuity liability not provided for": 255,
    "Disputed excise / customs / tax liabilities": 256,
    "Bank guarantee / Letter of credit outstanding": 257, "Other contingent liabilities": 258,
}
ALL_FIELDS = {**PNL, **BS}
ROW_TO_NAME = {v: k for k, v in ALL_FIELDS.items()}


def compare_cell(expected, actual, tol=0.02):
    e_empty = (expected is None or expected == ""
               or (isinstance(expected, (int, float)) and abs(expected) < 1e-9))
    a_empty = (actual is None or actual == ""
               or (isinstance(actual, (int, float)) and abs(actual) < 1e-9))
    if e_empty and a_empty:
        return ("BOTH_EMPTY", 0)
    if e_empty and not a_empty:
        return ("EXTRA", None)
    if not e_empty and a_empty:
        return ("MISSING", None)
    if not isinstance(expected, (int, float)) or not isinstance(actual, (int, float)):
        return ("NON_NUMERIC", None)
    if abs(expected) < 0.001:
        if abs(actual) < 0.001:
            return ("MATCH", 0)
        return ("MISMATCH", float("inf"))
    diff = abs(actual - expected) / abs(expected)
    if diff <= tol:
        return ("MATCH", diff)
    return ("MISMATCH", diff)


# ── Read all data-entry cells ──
COL_MAP = {2: "B(2021)", 3: "C(2022)", 4: "D(2023)"}
COL_LETTER = {2: "B", 3: "C", 4: "D"}
results = []
formula_eval_stats = {"evaluated": 0, "failed": 0, "not_needed": 0}

for fname, row in sorted(ALL_FIELDS.items(), key=lambda x: x[1]):
    for col in [2, 3, 4]:
        gt_val = ws_gt.cell_value(row - 1, col - 1)
        gen_cached = ws_gen_v.cell(row=row, column=col).value
        gen_formula_raw = ws_gen_f.cell(row=row, column=col).value

        if isinstance(gt_val, str) and gt_val.strip() == "":
            gt_val = None
        if isinstance(gen_cached, str) and gen_cached.strip() == "":
            gen_cached = None

        # Determine the actual generated value
        # Priority: cached value > formula evaluation
        gen_val = gen_cached
        formula_str = ""

        if gen_formula_raw is not None and isinstance(gen_formula_raw, str) and gen_formula_raw.startswith("="):
            formula_str = gen_formula_raw
            if gen_cached is None or gen_cached == "" or gen_cached == 0:
                # Try to evaluate the formula
                evaluated = eval_formula(gen_formula_raw)
                if evaluated is not None:
                    gen_val = evaluated
                    formula_eval_stats["evaluated"] += 1
                else:
                    formula_eval_stats["failed"] += 1
            else:
                formula_eval_stats["not_needed"] += 1
        elif gen_formula_raw is not None and gen_formula_raw != gen_cached:
            formula_str = str(gen_formula_raw)

        status, diff = compare_cell(gt_val, gen_val)

        results.append({
            "row": row, "field": fname, "fy": COL_MAP[col],
            "col_letter": COL_LETTER[col],
            "expected": gt_val, "got": gen_val,
            "cached": gen_cached, "formula": formula_str,
            "status": status, "diff": diff,
        })

# ── Summarize ──
gt_nonzero = [r for r in results if r["status"] in ("MATCH", "MISMATCH", "MISSING")]
matches = [r for r in results if r["status"] == "MATCH"]
mismatches = [r for r in results if r["status"] == "MISMATCH"]
missing = [r for r in results if r["status"] == "MISSING"]
extra = [r for r in results if r["status"] == "EXTRA"]
both_empty = [r for r in results if r["status"] == "BOTH_EMPTY"]
acc = len(matches) / len(gt_nonzero) * 100 if gt_nonzero else 0

print(f"\n=== FORMULA EVALUATION STATS ===")
print(f"  Formulas evaluated: {formula_eval_stats['evaluated']}")
print(f"  Formulas failed to eval: {formula_eval_stats['failed']}")
print(f"  Cached value used: {formula_eval_stats['not_needed']}")

print(f"\n=== DATA-ENTRY ACCURACY ===")
print(f"Total cells checked: {len(results)}")
print(f"GT has value: {len(gt_nonzero)}")
print(f"MATCH (within 2%): {len(matches)}")
print(f"MISMATCH: {len(mismatches)}")
print(f"MISSING: {len(missing)}")
print(f"EXTRA: {len(extra)}")
print(f"BOTH_EMPTY: {len(both_empty)}")
print(f"ACCURACY: {acc:.1f}% ({len(matches)}/{len(gt_nonzero)})")

print(f"\n=== MATCHING CELLS ({len(matches)}) ===")
for r in sorted(matches, key=lambda x: (x["row"], x["fy"])):
    exp = f"{r['expected']:.4f}" if isinstance(r["expected"], (int, float)) else str(r["expected"])
    got = f"{r['got']:.4f}" if isinstance(r["got"], (int, float)) else str(r["got"])
    d = f"{r['diff']*100:.1f}%" if r["diff"] is not None else ""
    f = r["formula"][:60] if r["formula"] else "(scalar)"
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | exp={exp:>10s} got={got:>10s} | {d:>5s} | {f}")

print(f"\n=== MISMATCHED CELLS ({len(mismatches)}) ===")
for r in sorted(mismatches, key=lambda x: (x["row"], x["fy"])):
    exp = f"{r['expected']:.4f}" if isinstance(r["expected"], (int, float)) else str(r["expected"])
    got_v = r["got"]
    got = f"{got_v:.4f}" if isinstance(got_v, (int, float)) else str(got_v)
    d = f"{r['diff']*100:.0f}%" if r["diff"] is not None and r["diff"] != float("inf") else "INF"
    f = r["formula"][:90] if r["formula"] else "(scalar)"
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | exp={exp:>10s} got={got:>12s} | diff={d:>8s} | {f}")

print(f"\n=== MISSING CELLS ({len(missing)}) ===")
missing_by_row = {}
for r in sorted(missing, key=lambda x: (x["row"], x["fy"])):
    key = (r["row"], r["field"])
    if key not in missing_by_row:
        missing_by_row[key] = {"fys": [], "expected": []}
    missing_by_row[key]["fys"].append(r["fy"])
    missing_by_row[key]["expected"].append(r["expected"])
    f = r["formula"]
    if f:
        missing_by_row[key]["formula"] = f

for (row, field), info in missing_by_row.items():
    fys = ", ".join(info["fys"])
    exp_range = f"{min(info['expected']):.2f}-{max(info['expected']):.2f}" if len(info["expected"]) > 1 else f"{info['expected'][0]:.4f}"
    note = ""
    if "formula" in info:
        note = f" [has formula: {info['formula'][:60]}]"
    print(f"  Row {row:3d} | {field[:40]:40s} | {fys:30s} | exp={exp_range}{note}")

print(f"\n=== EXTRA CELLS ({len(extra)}) ===")
for r in sorted(extra, key=lambda x: (x["row"], x["fy"])):
    got = f"{r['got']:.4f}" if isinstance(r["got"], (int, float)) else str(r["got"])
    f = r["formula"][:80] if r["formula"] else "(scalar)"
    print(f"  Row {r['row']:3d} | {r['field'][:40]:40s} | {r['fy']} | got={got} | {f}")

# ── Unit conversion check ──
print(f"\n=== UNIT CONVERSION CHECK (10 key rows) ===")
key_rows = [22, 42, 45, 83, 99, 116, 162, 206, 213, 242]
for row in key_rows:
    name = ROW_TO_NAME.get(row, "?")

    def get_gen_effective(r, c):
        cached = ws_gen_v.cell(row=r, column=c).value
        formula = ws_gen_f.cell(row=r, column=c).value
        if cached is not None and cached != "" and cached != 0:
            return cached
        if formula and isinstance(formula, str) and formula.startswith("="):
            ev = eval_formula(formula)
            return ev if ev is not None else f"[{formula[:40]}]"
        return cached

    gt_vals = [ws_gt.cell_value(row - 1, c - 1) for c in [2, 3, 4]]
    gen_vals = [get_gen_effective(row, c) for c in [2, 3, 4]]

    def fmt(v):
        if v is None or v == "" or v == 0:
            return "EMPTY"
        if isinstance(v, float):
            return f"{v:.2f}"
        return str(v)[:15]

    # Magnitude check
    gt_mags = [abs(v) if isinstance(v, (int, float)) and v != 0 else None for v in gt_vals]
    gen_mags = [abs(v) if isinstance(v, (int, float)) and v != 0 else None for v in gen_vals]
    mag_ok = "N/A"
    gt_present = [m for m in gt_mags if m is not None]
    gen_present = [m for m in gen_mags if m is not None]
    if gt_present and gen_present:
        ratio = max(gen_present) / min(gen_present) if min(gen_present) > 0.001 else 999
        gt_ratio = max(gt_present) / min(gt_present) if min(gt_present) > 0.001 else 999
        if ratio > 100:
            mag_ok = "NO (>100x spread)"
        elif ratio > 10:
            mag_ok = "WARN (>10x spread)"
        else:
            mag_ok = "OK"

    print(f"  Row {row:3d} {name[:30]:30s} | GT=[{fmt(gt_vals[0]):>10s}, {fmt(gt_vals[1]):>10s}, {fmt(gt_vals[2]):>10s}]"
          f" | Gen=[{fmt(gen_vals[0]):>10s}, {fmt(gen_vals[1]):>10s}, {fmt(gen_vals[2]):>10s}] | {mag_ok}")

    # Show formulas
    for ci, cl in [(2, "B"), (3, "C"), (4, "D")]:
        gf = ws_gen_f.cell(row=row, column=ci).value
        if gf and isinstance(gf, str) and gf.startswith("="):
            print(f"    {cl} formula: {gf[:120]}")

# ── Save raw JSON ──
output = {
    "total": len(results),
    "gt_nonzero": len(gt_nonzero),
    "matches": len(matches),
    "mismatches": len(mismatches),
    "missing": len(missing),
    "extra": len(extra),
    "both_empty": len(both_empty),
    "accuracy": round(acc, 1),
    "formula_eval_stats": formula_eval_stats,
    "match_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"], "got": r["got"], "diff": r["diff"],
         "formula": r["formula"]}
        for r in matches
    ],
    "mismatch_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"], "got": r["got"],
         "diff": r["diff"] if r["diff"] != float("inf") else 99999,
         "formula": r["formula"]}
        for r in mismatches
    ],
    "missing_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "expected": r["expected"], "formula": r["formula"]}
        for r in missing
    ],
    "extra_details": [
        {"row": r["row"], "field": r["field"], "fy": r["fy"],
         "got": r["got"], "formula": r["formula"]}
        for r in extra
    ],
}
out_path = os.path.join(BASE, "comparison-raw-eval.json")
with open(out_path, "w") as f:
    json.dump(output, f, indent=2, default=str)
print(f"\nRaw results saved to {out_path}")
