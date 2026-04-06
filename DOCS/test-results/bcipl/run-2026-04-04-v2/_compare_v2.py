"""V2 CMA Excel Comparison: Generated vs Ground Truth"""
import openpyxl
import xlrd
import os
import json

# --- FILE PATHS ---
gen_path = "C:/Users/ASHUTOSH/OneDrive/Desktop/CMA project -2/DOCS/test-results/bcipl/run-2026-04-04-v2/BCIPL_generated_CMA.xlsm"
gt_path = "C:/Users/ASHUTOSH/OneDrive/Desktop/FInancial Analysis/BCIPL/CMA BCIPL 12012024.xls"

assert os.path.exists(gen_path), f"Generated file not found: {gen_path}"
assert os.path.exists(gt_path), f"Ground truth file not found: {gt_path}"

print(f"Generated file size: {os.path.getsize(gen_path):,} bytes")
print(f"Ground truth file size: {os.path.getsize(gt_path):,} bytes")

# --- OPEN FILES ---
wb_gen_f = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=False)
wb_gen_v = openpyxl.load_workbook(gen_path, keep_vba=True, data_only=True)
wb_gt = xlrd.open_workbook(gt_path)

ws_gen_f = wb_gen_f["INPUT SHEET"]
ws_gen_v = wb_gen_v["INPUT SHEET"]
ws_gt = wb_gt.sheet_by_name("INPUT SHEET")

print(f"Generated sheets: {wb_gen_f.sheetnames}")
print(f"Ground truth sheets: {wb_gt.sheet_names()}")

# --- STRUCTURE CHECKS ---
structure = {}
structure["vba_present"] = wb_gen_f.vba_archive is not None
structure["client_name"] = ws_gen_v.cell(row=7, column=1).value
structure["years"] = [ws_gen_v.cell(row=8, column=c).value for c in [2,3,4]]
structure["nature"] = [ws_gen_v.cell(row=10, column=c).value for c in [2,3,4]]
structure["units"] = ws_gen_v.cell(row=13, column=2).value
structure["proj_years"] = [ws_gen_v.cell(row=8, column=c).value for c in [5,6,7,8]]

print("\n=== STRUCTURE ===")
for k, v in structure.items():
    print(f"  {k}: {v}")

# --- DATA-ENTRY ROWS ---
PNL_FIELD_TO_ROW = {
    "Domestic Sales": 22, "Export Sales": 23, "Less Excise Duty and Cess": 25,
    "Dividends received from Mutual Funds": 29, "Interest Received": 30,
    "Profit on sale of fixed assets / Investments": 31, "Gain on Exchange Fluctuations": 32,
    "Extraordinary income": 33, "Others (Non-Operating Income)": 34,
    "Raw Materials Consumed (Imported)": 41, "Raw Materials Consumed (Indigenous)": 42,
    "Stores and spares consumed (Imported)": 43, "Stores and spares consumed (Indigenous)": 44,
    "Wages": 45, "Processing / Job Work Charges": 46, "Freight and Transportation Charges": 47,
    "Power, Coal, Fuel and Water": 48, "Others (Manufacturing)": 49,
    "Repairs & Maintenance (Manufacturing)": 50, "Security Service Charges": 51,
    "Stock in process Opening Balance": 53, "Stock in process Closing Balance": 54,
    "Depreciation (Manufacturing)": 56, "Finished Goods Opening Balance": 58,
    "Finished Goods Closing Balance": 59, "Depreciation (CMA)": 63,
    "Other Manufacturing Exp (CMA)": 64, "Salary and staff expenses": 67,
    "Rent, Rates and Taxes": 68, "Bad Debts": 69,
    "Advertisements and Sales Promotions": 70, "Others (Admin)": 71,
    "Repairs & Maintenance (Admin)": 72, "Audit Fees & Directors Remuneration": 73,
    "Miscellaneous Expenses written off": 75, "Deferred Revenue Expenditures": 76,
    "Other Amortisations": 77, "Interest on Fixed Loans / Term loans": 83,
    "Interest on Working capital loans": 84, "Bank Charges": 85,
    "Loss on sale of fixed assets / Investments": 89, "Sundry Balances Written off": 90,
    "Loss on Exchange Fluctuations": 91, "Extraordinary losses": 92,
    "Others (Non-Operating Expenses)": 93, "Income Tax provision": 99,
    "Deferred Tax Liability (P&L)": 100, "Deferred Tax Asset (P&L)": 101,
    "Brought forward from previous year": 106, "Dividend": 107,
    "Other Appropriation of profit": 108,
}

BS_FIELD_TO_ROW = {
    "Issued, Subscribed and Paid up": 116, "Share Application Money": 117,
    "General Reserve": 121, "Balance transferred from profit and loss a/c": 122,
    "Share Premium A/c": 123, "Revaluation Reserve": 124, "Other Reserve": 125,
    "Working Capital Bank Finance - Bank 1": 131, "Working Capital Bank Finance - Bank 2": 132,
    "Term Loan Repayable in next one year": 136, "Term Loan Balance Repayable after one year": 137,
    "Debentures Repayable in next one year": 140, "Debentures Balance Repayable after one year": 141,
    "Preference Shares Repayable in next one year": 144, "Preference Shares Balance Repayable after one year": 145,
    "Other Debts Repayable in Next One year": 148, "Balance Other Debts": 149,
    "Unsecured Loans - Quasi Equity": 152, "Unsecured Loans - Long Term Debt": 153,
    "Unsecured Loans - Short Term Debt": 154, "Deferred tax liability (BS)": 159,
    "Gross Block": 162, "Less Accumulated Depreciation": 163, "Capital Work in Progress": 165,
    "Patents / goodwill / copyrights etc": 169, "Misc Expenditure (to the extent not w/o)": 170,
    "Deferred Tax Asset (BS)": 171, "Other Intangible assets": 172,
    "Additions to Fixed Assets": 175, "Sale of Fixed assets WDV": 176,
    "Profit on sale of Fixed assets (BS)": 177, "Loss on sale of Fixed assets (BS)": 178,
    "Investment in Govt. Securities (Current)": 182, "Investment in Govt. Securities (Non Current)": 183,
    "Other current investments": 185, "Other non current investments": 186,
    "Investment in group companies / subsidiaries": 188,
    "Raw Material Imported": 193, "Raw Material Indigenous": 194,
    "Stores and Spares Imported": 197, "Stores and Spares Indigenous": 198,
    "Stocks-in-process": 200, "Finished Goods": 201,
    "Domestic Receivables": 206, "Export Receivables": 207, "Debtors more than 6 months": 208,
    "Cash on Hand": 212, "Bank Balances": 213,
    "Fixed Deposit under lien": 214, "Other Fixed Deposits": 215,
    "Advances recoverable in cash or in kind": 219, "Advances to suppliers of raw materials": 220,
    "Advance Income Tax": 221, "Prepaid Expenses": 222,
    "Other Advances / current asset": 223, "Advances to group / subsidiaries companies": 224,
    "Exposure in group companies - Investments": 229, "Exposure in group companies - Advances": 230,
    "Debtors more than six months (Non-Current)": 232, "Investments (Non-Current)": 233,
    "Fixed Deposits (Non Current)": 234, "Dues from directors / partners / promoters": 235,
    "Advances to suppliers of capital goods": 236, "Security deposits with government departments": 237,
    "Other non current assets": 238,
    "Sundry Creditors for goods": 242, "Advance received from customers": 243,
    "Provision for Taxation": 244, "Dividend payable": 245,
    "Other statutory liabilities (due within 1 year)": 246,
    "Interest Accrued but not due": 247, "Interest Accrued and due": 248,
    "Creditors for Expenses": 249, "Other current liabilities": 250,
    "Arrears of cumulative dividends": 254, "Gratuity liability not provided for": 255,
    "Disputed excise / customs / tax liabilities": 256,
    "Bank guarantee / Letter of credit outstanding": 257, "Other contingent liabilities": 258,
}

ALL_FIELD_TO_ROW = {**PNL_FIELD_TO_ROW, **BS_FIELD_TO_ROW}
ROW_TO_FIELD = {v: k for k, v in ALL_FIELD_TO_ROW.items()}

# --- COMPARISON LOGIC ---
FY_COLS = {"B(2021)": 2, "C(2022)": 3, "D(2023)": 4}

def get_gt_value(row, col):
    try:
        v = ws_gt.cell_value(row - 1, col - 1)
        if v == "" or v is None:
            return None
        return float(v) if isinstance(v, (int, float)) else None
    except:
        return None

def get_gen_value(row, col):
    try:
        v = ws_gen_v.cell(row=row, column=col).value
        if v is None or v == "":
            return None
        return float(v)
    except:
        return None

def get_gen_formula(row, col):
    try:
        v = ws_gen_f.cell(row=row, column=col).value
        if v is None:
            return None
        return str(v)
    except:
        return None

def compare_cell(expected, actual, tolerance=0.02):
    if (expected is None or expected == 0) and (actual is None or actual == 0):
        return "BOTH_EMPTY", 0
    if expected is None or expected == 0:
        if actual is not None and actual != 0:
            return "EXTRA", None
        return "BOTH_EMPTY", 0
    if actual is None or actual == 0:
        return "MISSING", None
    if abs(expected) < 0.001 and abs(actual) < 0.001:
        return "MATCH", 0
    if abs(expected) < 0.001:
        return "MISMATCH", float("inf")
    diff_pct = abs(actual - expected) / abs(expected)
    if diff_pct <= tolerance:
        return "MATCH", diff_pct
    return "MISMATCH", diff_pct

# --- RUN COMPARISON ---
results = []
for field_name, row in sorted(ALL_FIELD_TO_ROW.items(), key=lambda x: x[1]):
    for fy_label, col in FY_COLS.items():
        gt_val = get_gt_value(row, col)
        gen_val = get_gen_value(row, col)
        gen_formula = get_gen_formula(row, col)
        status, diff = compare_cell(gt_val, gen_val)
        results.append({
            "row": row,
            "field": field_name,
            "fy": fy_label,
            "expected": gt_val,
            "got": gen_val,
            "formula": gen_formula,
            "status": status,
            "diff_pct": diff,
        })

# --- SUMMARIZE ---
gt_has_value = [r for r in results if r["expected"] is not None and r["expected"] != 0]
matches = [r for r in results if r["status"] == "MATCH"]
mismatches = [r for r in results if r["status"] == "MISMATCH"]
missing = [r for r in results if r["status"] == "MISSING"]
extra = [r for r in results if r["status"] == "EXTRA"]
both_empty = [r for r in results if r["status"] == "BOTH_EMPTY"]

print(f"\n=== DATA-ENTRY ACCURACY SUMMARY ===")
print(f"Total data-entry cells checked: {len(results)}")
print(f"GT has value (non-empty, non-zero): {len(gt_has_value)}")
print(f"Matching (within 2%): {len(matches)}")
print(f"Mismatched: {len(mismatches)}")
print(f"Missing (GT has value, gen empty): {len(missing)}")
print(f"Extra (gen has value, GT empty): {len(extra)}")
print(f"Both empty: {len(both_empty)}")
accuracy = len(matches) / len(gt_has_value) * 100 if gt_has_value else 0
print(f"Data-entry accuracy: {accuracy:.1f}% ({len(matches)}/{len(gt_has_value)})")

# --- DUMP RAW JSON ---
out_path = "C:/Users/ASHUTOSH/OneDrive/Desktop/CMA project -2/DOCS/test-results/bcipl/run-2026-04-04-v2/_comparison_raw.json"
with open(out_path, "w") as f:
    json.dump(results, f, indent=2, default=str)
print(f"\nRaw results saved to {out_path}")

# --- PRINT MATCHES ---
print(f"\n=== MATCHING CELLS ({len(matches)}) ===")
for r in sorted(matches, key=lambda x: (x["row"], x["fy"])):
    diff_s = f"{r['diff_pct']*100:.1f}%" if r["diff_pct"] else "0.0%"
    print(f"  Row {r['row']:3d} ({r['field'][:35]:35s}) {r['fy']}: expected={r['expected']:.4f}, got={r['got']:.4f}, diff={diff_s}")

# --- PRINT MISMATCHES ---
print(f"\n=== MISMATCHED CELLS ({len(mismatches)}) ===")
for r in sorted(mismatches, key=lambda x: (x["row"], x["fy"])):
    diff_str = f"{r['diff_pct']*100:.0f}%" if r["diff_pct"] and r["diff_pct"] != float("inf") else "inf"
    formula_str = (r["formula"] or "")[:60]
    print(f"  Row {r['row']:3d} ({r['field'][:30]:30s}) {r['fy']}: expected={r['expected']:.4f}, got={r['got']:.4f}, diff={diff_str}, formula={formula_str}")

# --- PRINT MISSING ---
print(f"\n=== MISSING CELLS ({len(missing)}) ===")
for r in sorted(missing, key=lambda x: (x["row"], x["fy"])):
    print(f"  Row {r['row']:3d} ({r['field'][:35]:35s}) {r['fy']}: expected={r['expected']:.4f}")

# --- PRINT EXTRA ---
print(f"\n=== EXTRA CELLS ({len(extra)}) ===")
for r in sorted(extra, key=lambda x: (x["row"], x["fy"])):
    formula_str = (r["formula"] or "")[:60]
    print(f"  Row {r['row']:3d} ({r['field'][:35]:35s}) {r['fy']}: got={r['got']:.4f}, formula={formula_str}")

# --- KEY GROUND TRUTH SANITY CHECK ---
print(f"\n=== KEY GROUND TRUTH VALUES (sanity) ===")
key_rows = [22, 23, 42, 45, 83, 116, 162, 206, 242]
for row in key_rows:
    name = ROW_TO_FIELD.get(row, "?")
    vals = []
    for col in [2, 3, 4]:
        v = get_gt_value(row, col)
        vals.append(f"{v:.2f}" if v else "empty")
    gen_vals = []
    for col in [2, 3, 4]:
        v = get_gen_value(row, col)
        gen_vals.append(f"{v:.2f}" if v else "empty")
    print(f"  Row {row:3d} ({name[:30]:30s}) GT: {vals[0]:>10s} {vals[1]:>10s} {vals[2]:>10s}  |  GEN: {gen_vals[0]:>10s} {gen_vals[1]:>10s} {gen_vals[2]:>10s}")

# --- UNIT CONVERSION CHECK ---
print(f"\n=== UNIT CONVERSION CHECK (10 key rows) ===")
check_rows = [22, 42, 45, 83, 99, 116, 162, 206, 213, 242]
for row in check_rows:
    name = ROW_TO_FIELD.get(row, "?")
    gt_b = get_gt_value(row, 2)
    gt_c = get_gt_value(row, 3)
    gt_d = get_gt_value(row, 4)
    gen_b = get_gen_value(row, 2)
    gen_c = get_gen_value(row, 3)
    gen_d = get_gen_value(row, 4)
    # Check if generated values are same order of magnitude across FYs
    gen_vals = [v for v in [gen_b, gen_c, gen_d] if v and abs(v) > 0.001]
    if len(gen_vals) >= 2:
        max_v = max(abs(v) for v in gen_vals)
        min_v = min(abs(v) for v in gen_vals)
        ratio = max_v / min_v if min_v > 0 else float("inf")
        mag_ok = ratio < 50
    else:
        mag_ok = None
    gt_s = lambda v: f"{v:.4f}" if v else "empty"
    gen_s = lambda v: f"{v:.4f}" if v else "empty"
    mag_str = "OK" if mag_ok is True else ("BAD" if mag_ok is False else "N/A")
    print(f"  Row {row:3d} ({name[:25]:25s}) GT: {gt_s(gt_b):>10s} {gt_s(gt_c):>10s} {gt_s(gt_d):>10s}  GEN: {gen_s(gen_b):>10s} {gen_s(gen_c):>10s} {gen_s(gen_d):>10s}  Mag: {mag_str}")
