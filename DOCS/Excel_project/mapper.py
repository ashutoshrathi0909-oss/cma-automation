import openpyxl
from openpyxl.utils import get_column_letter
import os

# --- CONFIGURATION ---
# REPLACE 'MyData.xlsm' WITH YOUR EXACT FILE NAME
INPUT_FILE_NAME = 'MyData.xlsm' 
OUTPUT_FILE_NAME = 'AI_Explanation.txt'
# ---------------------

def analyze_excel_for_ai():
    print(f"Reading {INPUT_FILE_NAME}... please wait.")
    
    if not os.path.exists(INPUT_FILE_NAME):
        print(f"ERROR: Could not find '{INPUT_FILE_NAME}' in this folder.")
        return

    try:
        # Load the file. data_only=False ensures we get FORMULAS, not just numbers.
        wb = openpyxl.load_workbook(INPUT_FILE_NAME, data_only=False, keep_vba=True)
        
        with open(OUTPUT_FILE_NAME, "w", encoding="utf-8") as f:
            f.write(f"# EXCEL DOCUMENTATION FOR AI\n")
            f.write(f"File Name: {INPUT_FILE_NAME}\n")
            f.write("GOAL: Explain the logic, data flow, and schema of this workbook.\n\n")
            
            # 1. MACRO DETECTION
            if wb.vba_archive:
                f.write("## MACROS / VBA\n")
                f.write("⚠️ This workbook contains Macros (VBA). The Python script cannot read VBA code text directly.\n")
                f.write("USER INSTRUCTION: Please copy the VBA code manually from the Developer tab and paste it after this report.\n\n")

            # 2. SHEET ANALYSIS
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                f.write(f"--- SHEET: {sheet_name} ---\n")
                
                # Get Headers (Row 1)
                headers = []
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=1, column=col)
                    if cell.value:
                        headers.append(f"Col {get_column_letter(col)}: '{cell.value}'")
                
                f.write("### Columns:\n")
                f.write(", ".join(headers) + "\n\n")

                # Get Formulas (Scan first 20 rows to find logic)
                f.write("### Formulas & Logic:\n")
                found_formulas = {} # Store unique formulas to avoid repetition
                
                # Check rows 2 to 20
                for row_idx in range(2, 21): 
                    for col_idx in range(1, ws.max_column + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        if cell.data_type == 'f': # 'f' means Formula
                            col_letter = get_column_letter(col_idx)
                            # We create a generic pattern so AI understands the rule
                            # e.g. changes =SUM(A2:B2) to =SUM(A[ROW]:B[ROW])
                            formula_str = str(cell.value).replace(str(row_idx), "[ROW]")
                            
                            if col_letter not in found_formulas:
                                found_formulas[col_letter] = formula_str
                
                if found_formulas:
                    for col, formula in found_formulas.items():
                        f.write(f"- Column {col} uses logic: {formula}\n")
                else:
                    f.write("- No formulas found in top 20 rows (Raw Data).\n")
                f.write("\n")

        print(f"DONE! Created '{OUTPUT_FILE_NAME}'.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    analyze_excel_for_ai()