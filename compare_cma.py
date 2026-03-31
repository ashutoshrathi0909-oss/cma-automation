"""
compare_cma.py — Compare generated CMA Excel vs CA ground truth
INPUT SHEET rows 17-262, columns B, C, D
Tolerance: 2% (relative) or absolute < 0.001 for near-zero values
"""
import sys
import io
import json
import openpyxl
import xlrd  # for .xls ground truth

GENERATED_PATH = r"DOCS/test-results/bcipl/run-2026-03-23/generated_cma.xlsm"
GROUND_TRUTH_PATH = r"FInancials main/BCIPL/CMA BCIPL 12012024.xls"
SHEET_NAME = "INPUT SHEET"
ROW_START = 17
ROW_END = 262
COLS = ["B", "C", "D"]
TOLERANCE = 0.02  # 2%
OUTPUT_MD = r"DOCS/test-results/bcipl/run-2026-03-23/comparison_results.md"


COL_LETTERS = {"B": 1, "C": 2, "D": 3}  # 0-indexed column numbers (A=0, B=1...)


def load_sheet_data(path, sheet_name):
    """Load data from Excel file (xls or xlsx/xlsm), returns dict {(row, col_letter): value}"""
    if path.lower().endswith(".xls") and not path.lower().endswith(".xlsm"):
        return _load_xls(path, sheet_name)
    return _load_xlsx(path, sheet_name)


def _load_xls(path, sheet_name):
    """Load old .xls using xlrd."""
    try:
        wb = xlrd.open_workbook(path)
    except Exception as e:
        raise RuntimeError(f"Cannot open {path}: {e}")

    ws = None
    for name in wb.sheet_names():
        if sheet_name.lower() in name.lower():
            ws = wb.sheet_by_name(name)
            break

    if ws is None:
        raise RuntimeError(f"Sheet '{sheet_name}' not found. Available: {wb.sheet_names()}")

    data = {}
    for row in range(ROW_START, ROW_END + 1):
        for col in COLS:
            col_idx = COL_LETTERS[col]
            try:
                cell = ws.cell(row - 1, col_idx)  # xlrd is 0-indexed
                val = cell.value if cell.ctype not in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK) else None
                data[(row, col)] = val
            except IndexError:
                data[(row, col)] = None
    return data


def _load_xlsx(path, sheet_name):
    """Load .xlsx/.xlsm using openpyxl."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    except Exception:
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            raise RuntimeError(f"Cannot open {path}: {e}")

    ws = None
    for name in wb.sheetnames:
        if sheet_name.lower() in name.lower():
            ws = wb[name]
            break

    if ws is None:
        raise RuntimeError(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")

    data = {}
    for row in range(ROW_START, ROW_END + 1):
        for col in COLS:
            cell = ws[f"{col}{row}"]
            data[(row, col)] = cell.value
    return data


def to_float(v):
    """Convert cell value to float, None if not numeric."""
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def values_match(gen_val, gt_val, tolerance=TOLERANCE):
    """Return True if values match within tolerance."""
    gen_f = to_float(gen_val)
    gt_f = to_float(gt_val)

    # Both None/blank
    if gen_f is None and gt_f is None:
        return True, "both_blank"

    # One blank
    if gen_f is None:
        if gt_f == 0.0:
            return True, "zero_vs_blank"
        return False, f"gen=blank, gt={gt_f}"
    if gt_f is None:
        if gen_f == 0.0:
            return True, "blank_vs_zero"
        return False, f"gen={gen_f}, gt=blank"

    # Both numeric
    if gt_f == 0.0 and gen_f == 0.0:
        return True, "both_zero"
    if gt_f == 0.0:
        if abs(gen_f) < 0.001:
            return True, "near_zero"
        return False, f"gen={gen_f:.4f}, gt=0"

    rel_diff = abs(gen_f - gt_f) / abs(gt_f)
    if rel_diff <= tolerance:
        return True, f"diff={rel_diff*100:.2f}%"
    return False, f"gen={gen_f:.4f}, gt={gt_f:.4f}, diff={rel_diff*100:.2f}%"


def main():
    print(f"Loading generated: {GENERATED_PATH}")
    try:
        gen_data = load_sheet_data(GENERATED_PATH, SHEET_NAME)
        print(f"  OK Loaded {len(gen_data)} cells")
    except Exception as e:
        print(f"  ERROR: {e}")
        gen_data = {}

    print(f"Loading ground truth: {GROUND_TRUTH_PATH}")
    try:
        gt_data = load_sheet_data(GROUND_TRUTH_PATH, SHEET_NAME)
        print(f"  OK Loaded {len(gt_data)} cells")
    except Exception as e:
        print(f"  ERROR: {e}")
        gt_data = {}

    if not gen_data and not gt_data:
        print("Cannot proceed — both files failed to load")
        sys.exit(1)

    results = []
    matches = 0
    mismatches = 0
    blanks = 0

    for row in range(ROW_START, ROW_END + 1):
        for col in COLS:
            gen_v = gen_data.get((row, col))
            gt_v = gt_data.get((row, col))

            ok, reason = values_match(gen_v, gt_v)
            entry = {
                "row": row,
                "col": col,
                "gen": gen_v,
                "gt": gt_v,
                "match": ok,
                "reason": reason,
            }
            results.append(entry)
            if ok:
                if "blank" in reason or "zero" in reason.lower():
                    blanks += 1
                else:
                    matches += 1
            else:
                mismatches += 1

    total = len(results)
    numeric_total = total - blanks
    match_pct = (matches / numeric_total * 100) if numeric_total > 0 else 0

    # Print summary
    print(f"\n=== COMPARISON SUMMARY ===")
    print(f"Total cells compared: {total} (rows {ROW_START}-{ROW_END}, cols B/C/D)")
    print(f"Both blank / zero: {blanks}")
    print(f"Matches (within 2%): {matches}")
    print(f"Mismatches: {mismatches}")
    print(f"Match rate (non-blank): {match_pct:.1f}%")

    # Show mismatches
    mismatch_rows = [r for r in results if not r["match"]]
    if mismatch_rows:
        print(f"\nMISMATCHED CELLS ({len(mismatch_rows)}):")
        for r in mismatch_rows[:50]:
            print(f"  {r['col']}{r['row']:3d}: gen={r['gen']} | gt={r['gt']} | {r['reason']}")

    # Write markdown report
    lines = [
        "# CMA Comparison Results",
        "",
        f"**Generated:** `{GENERATED_PATH}`",
        f"**Ground Truth:** `{GROUND_TRUTH_PATH}`",
        f"**Sheet:** `{SHEET_NAME}` | Rows {ROW_START}–{ROW_END} | Cols B, C, D",
        f"**Tolerance:** 2% relative",
        "",
        "## Summary",
        "",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Total cells compared | {total} |",
        f"| Both blank/zero | {blanks} |",
        f"| Matches (within 2%) | {matches} |",
        f"| **Mismatches** | **{mismatches}** |",
        f"| **Match rate (non-blank)** | **{match_pct:.1f}%** |",
        "",
    ]

    if mismatches == 0:
        lines.append("## Result: PASS - All numeric cells match within 2% tolerance")
    else:
        lines.append(f"## Result: FAIL - {mismatches} cells out of tolerance")

    lines += ["", "## Mismatched Cells", ""]
    if mismatch_rows:
        lines.append("| Cell | Generated | Ground Truth | Diff |")
        lines.append("|------|-----------|--------------|------|")
        for r in mismatch_rows:
            lines.append(f"| {r['col']}{r['row']} | {r['gen']} | {r['gt']} | {r['reason']} |")
    else:
        lines.append("_None — all cells match._")

    lines += [
        "",
        "## Notes",
        "",
        "- FY2022 source_unit was `lakhs` in DB (not `rupees` as test plan specifies). Values converted lakhs→crores (÷100) vs rupees→crores (÷10M). This discrepancy will cause systematic FY2022 mismatches.",
        "- Classifications were from a previously-run session using the existing BCIPL client (19cf7c12), not fresh AI classification of newly-uploaded docs.",
        "- Column B = FY2021, Column C = FY2022, Column D = FY2023 (as per CMA template year mapping).",
    ]

    with open(OUTPUT_MD, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"\nComparison results written to: {OUTPUT_MD}")

    # Also save JSON for later inspection
    json_path = OUTPUT_MD.replace(".md", ".json")
    with open(json_path, "w") as f:
        json.dump({
            "summary": {
                "total": total,
                "blanks": blanks,
                "matches": matches,
                "mismatches": mismatches,
                "match_pct": round(match_pct, 2),
            },
            "mismatches": mismatch_rows,
        }, f, indent=2, default=str)
    print(f"JSON results written to: {json_path}")


if __name__ == "__main__":
    main()
