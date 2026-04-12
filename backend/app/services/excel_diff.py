"""Excel diff service — compares AI-generated vs father-corrected CMA files.

Pure Python, zero API calls. Produces a list of CellDiff records
identifying every cell where the father's version differs from AI's.
"""

from __future__ import annotations

from dataclasses import dataclass

import openpyxl
from openpyxl.utils import get_column_letter

INPUT_SHEET_NAME = "INPUT SHEET"
# CMA data region: rows 11 (first data row) through 258 (last BS asset row)
_DATA_ROW_START = 11
_DATA_ROW_END = 258
# Columns B through H (indices 2-8) hold financial year data
_DATA_COL_START = 2
_DATA_COL_END = 8

_NUMERIC_TOLERANCE = 0.01


@dataclass
class CellDiff:
    """One cell that differs between AI-generated and father-corrected CMA."""
    cma_row: int
    cma_column: str
    ai_value: float | None
    father_value: float | None


def _is_formula(value) -> bool:
    """Return True if value looks like an Excel formula."""
    return isinstance(value, str) and value.startswith("=")


def _numeric_equal(a, b) -> bool:
    """Compare two values with tolerance for floating point."""
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= _NUMERIC_TOLERANCE
    except (TypeError, ValueError):
        return a == b


def diff_cma_files(ai_path: str, father_path: str) -> list[CellDiff]:
    """Compare two CMA Excel files and return a list of cell differences.

    Parameters
    ----------
    ai_path     — path to the AI-generated CMA file
    father_path — path to the father-corrected CMA file

    Returns
    -------
    list[CellDiff] — one entry per cell that differs (formulas excluded)
    """
    ai_wb = openpyxl.load_workbook(ai_path, data_only=True)
    father_wb = openpyxl.load_workbook(father_path, data_only=True)

    ai_ws = ai_wb[INPUT_SHEET_NAME]
    father_ws = father_wb[INPUT_SHEET_NAME]

    diffs: list[CellDiff] = []

    for row in range(_DATA_ROW_START, _DATA_ROW_END + 1):
        for col in range(_DATA_COL_START, _DATA_COL_END + 1):
            ai_cell = ai_ws.cell(row=row, column=col).value
            father_cell = father_ws.cell(row=row, column=col).value

            # Skip formula cells in either file
            if _is_formula(ai_cell) or _is_formula(father_cell):
                continue

            # Skip if both are empty/None
            if ai_cell is None and father_cell is None:
                continue

            # Skip if numerically equal within tolerance
            if _numeric_equal(ai_cell, father_cell):
                continue

            col_letter = get_column_letter(col)
            diffs.append(CellDiff(
                cma_row=row,
                cma_column=col_letter,
                ai_value=ai_cell,
                father_value=father_cell,
            ))

    ai_wb.close()
    father_wb.close()

    return diffs
