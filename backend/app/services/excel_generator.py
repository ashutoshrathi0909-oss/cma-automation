"""CMA Excel generation service.

Fills the CMA.xlsm INPUT SHEET with classified financial data,
uploads the result to Supabase Storage, logs to audit trail,
and cleans up temp files.

Design
------
fill_workbook(ws, client_name, docs, cell_data)
    Pure transform: writes to an already-open worksheet.
    No DB calls, no file I/O — fully unit-testable.

generate(report_id, user_id) -> str
    Full pipeline: fetch data from DB, open template, fill,
    save temp file, upload, audit, cleanup.
    Returns the Supabase Storage path of the uploaded file.
"""

from __future__ import annotations

import logging
import math
import os
import re
import tempfile

import openpyxl
from openpyxl.utils import column_index_from_string

from app.mappings.cma_field_rows import ALL_FIELD_TO_ROW
from app.mappings.year_columns import build_year_map
from app.services.audit_service import log_action

logger = logging.getLogger(__name__)

_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$"
)

# Patterns that indicate a REAL formula (references other cells, sheets, or functions).
# Placeholder formulas like =0 or =0.00 do NOT match and are safe to overwrite.
_REAL_FORMULA_RE = re.compile(
    r"[A-Za-z][A-Za-z0-9]*\("   # function call: SUM(, IF(, VLOOKUP(
    r"|[A-Za-z]{1,3}\d+"        # cell reference: A1, BC123
    r"|!"                        # sheet reference: 'Sheet1'!A1
    r"|:"                        # range operator: A1:A10
)

def _format_number(v: float) -> str:
    """Format a number for an Excel formula: integer if whole, else 2 decimals."""
    if math.isnan(v) or math.isinf(v):
        return "0"
    rounded = round(v, 2)
    if rounded == int(rounded):
        return str(int(rounded))
    return f"{rounded:.2f}"


def _build_formula(values: list[float]) -> str:
    """Build an Excel formula showing individual values instead of their sum.

    Examples:
        [50000, 30000, -5000] → "=50000+30000-5000"
        [12000]               → "=12000"  (single value, still a formula for consistency)
    """
    if not values:
        return "=0"
    parts = []
    for i, v in enumerate(values):
        formatted = _format_number(v)
        if i == 0:
            parts.append(formatted)
        elif v < 0:
            parts.append(formatted)  # negative sign already included
        else:
            parts.append(f"+{formatted}")
    return "=" + "".join(parts)


DEFAULT_TEMPLATE_PATH = "/app/DOCS/CMA.xlsm"
STORAGE_BUCKET = "generated"
INPUT_SHEET_NAME = "INPUT SHEET"

# Header row constants (INPUT SHEET layout)
_ROW_CLIENT_NAME = 7
_ROW_YEAR = 8          # "Year" row — 2023, 2024, etc.
_ROW_MONTHS = 9        # "Number of months" — always 12, don't overwrite
_ROW_NATURE = 10       # "Nature of Financials" — Audited / Provisionals / Projected

# Unit conversion: source documents may be in full Rupees, Rs.'000, or Lakhs.
# CMA output may be in Lakhs or Crores depending on the CA's preference.
#
# Common divisors:
#   Full Rupees -> Lakhs:  100,000
#   Full Rupees -> Crores: 10,000,000
#   Rs.'000     -> Lakhs:  100
#   Rs.'000     -> Crores: 10,000
#   Lakhs       -> Lakhs:  1 (no conversion)
#   Lakhs       -> Crores: 100
#
# Set to 1 for no conversion.  Should be set per-report based on source/target units.
DEFAULT_UNIT_DIVISOR = 1  # No default assumption — must be set per report

# Maps unit name → value in Rupees (how many Rupees is 1 of this unit?)
_UNIT_IN_RUPEES: dict[str, float] = {
    "rupees": 1,
    "thousands": 1_000,
    "lakhs": 100_000,
    "crores": 10_000_000,
}


def compute_unit_divisor(source_unit: str, output_unit: str) -> float:
    """Return the number to divide source amounts by to get output amounts.

    Examples:
        compute_unit_divisor("rupees", "crores")   → 10,000,000
        compute_unit_divisor("thousands", "lakhs")  → 100
        compute_unit_divisor("lakhs", "lakhs")      → 1
    """
    src = _UNIT_IN_RUPEES.get(source_unit, 1)
    out = _UNIT_IN_RUPEES.get(output_unit, 100_000)  # default: lakhs
    if out == 0:
        return 1
    return out / src


class ExcelGenerator:
    """Generates the CMA Excel file from reviewed classifications."""

    def __init__(self, service, template_path: str | None = None) -> None:
        self.service = service
        self.template_path = template_path or DEFAULT_TEMPLATE_PATH

    # ── Public API ────────────────────────────────────────────────────────

    def fill_workbook(
        self,
        ws,
        client_name: str,
        docs: list[dict],
        cell_data: list[dict],
        unit_divisor: float = 1,
        doc_divisors: dict[str, float] | None = None,
    ) -> list[dict]:
        """Pure transform: fill *ws* (an open worksheet) with CMA data.

        Parameters
        ----------
        ws            — openpyxl Worksheet (INPUT SHEET)
        client_name   — client name for the header row
        docs          — list of {"financial_year": int, "nature": str}
        cell_data     — list of {"cma_field_name": str, "financial_year": int, "amount": float,
                          "document_id": str (optional)}
        unit_divisor  — DEPRECATED: single divisor for all items. Use doc_divisors instead.
        doc_divisors  — {document_id: divisor} for per-document unit conversion.
                        When provided, each item uses its own document's divisor.
        """
        # Build year→column mapping dynamically from the documents
        years = [d["financial_year"] for d in docs if d.get("financial_year")]
        year_map = build_year_map(years)
        logger.info("Dynamic year mapping: %s (base=%s)", year_map, min(years) if years else "N/A")

        self._fill_headers(ws, client_name, docs, year_map)
        return self._fill_data_cells(ws, cell_data, year_map=year_map, unit_divisor=unit_divisor, doc_divisors=doc_divisors)

    def generate(self, report_id: str, user_id: str) -> str:
        """Full pipeline: fetch → fill → save → upload → audit → cleanup.

        Returns
        -------
        str — Supabase Storage path of the uploaded .xlsm file
        """
        logger.info("ExcelGenerator.generate started for report_id=%s", report_id)

        # 1. Fetch everything needed
        report = self._fetch_report(report_id)
        client_name = self._fetch_client_name(report["client_id"])
        doc_ids: list[str] = report.get("document_ids") or []
        docs = self._fetch_documents(doc_ids)
        doc_years = {d["id"]: d.get("financial_year") for d in docs}
        cell_data = self._fetch_classified_data(doc_ids, doc_years)

        # 2. Open template (keep_vba=True — MANDATORY for macros)
        wb = openpyxl.load_workbook(self.template_path, keep_vba=True)
        ws = wb[INPUT_SHEET_NAME]

        # 3. Compute per-document unit divisors (each doc may have different source_unit)
        cma_output_unit = report.get("cma_output_unit") or "lakhs"
        doc_units = {d["id"]: d.get("source_unit") or "rupees" for d in docs}
        doc_divisors = {
            doc_id: compute_unit_divisor(src_unit, cma_output_unit)
            for doc_id, src_unit in doc_units.items()
        }
        logger.info(
            "Per-document unit conversion (output=%s): %s",
            cma_output_unit,
            {doc_id: f"{doc_units[doc_id]}→{cma_output_unit} (÷{div})" for doc_id, div in doc_divisors.items()},
        )

        # 4. Fill worksheet
        provenance = self.fill_workbook(ws, client_name, docs, cell_data, doc_divisors=doc_divisors)

        # 5. Insert provenance records
        _BATCH_SIZE = 500
        if provenance:
            for i in range(0, len(provenance), _BATCH_SIZE):
                batch = provenance[i : i + _BATCH_SIZE]
                records = [
                    {**p, "cma_report_id": report_id}
                    for p in batch
                ]
                try:
                    self.service.table("cell_provenance").insert(records).execute()
                except Exception as exc:
                    logger.error("Provenance insert failed (batch %d): %s", i // _BATCH_SIZE, exc)

        # 6. Save, upload, cleanup
        storage_path = self._save_upload_cleanup(wb, report_id, user_id)

        logger.info(
            "ExcelGenerator.generate complete: report_id=%s path=%s",
            report_id,
            storage_path,
        )
        return storage_path

    # ── Private: fill helpers ─────────────────────────────────────────────

    def _fill_headers(self, ws, client_name: str, docs: list[dict], year_map: dict[int, str]) -> None:
        ws.cell(row=_ROW_CLIENT_NAME, column=1).value = client_name

        # Build a lookup: year → nature from docs
        year_nature = {d["financial_year"]: d.get("nature", "") for d in docs if d.get("financial_year")}

        if not year_nature:
            return

        base_year = min(year_nature)
        num_historical = len(year_nature)

        # Rewrite ALL year cells in Row 8 (columns B through H = indices 2-8)
        # Historical columns get the actual years; projection columns continue the sequence
        for col_idx in range(2, 9):  # B(2) through H(8)
            year_for_col = base_year + (col_idx - 2)
            ws.cell(row=_ROW_YEAR, column=col_idx).value = year_for_col

            # Set nature: use doc's nature for historical years, "Projected" for the rest
            if year_for_col in year_nature:
                ws.cell(row=_ROW_NATURE, column=col_idx).value = year_nature[year_for_col]
            else:
                ws.cell(row=_ROW_NATURE, column=col_idx).value = "Projected"

        logger.info(
            "Header years: base=%d, historical=%d years, cols B-H = %d-%d",
            base_year, num_historical, base_year, base_year + 6,
        )

    def _fill_data_cells(
        self,
        ws,
        cell_data: list[dict],
        *,
        year_map: dict[int, str],
        unit_divisor: float = 1,
        doc_divisors: dict[str, float] | None = None,
    ) -> list[dict]:
        """Accumulate amounts by (row, col), apply per-document unit conversion, then write once per cell.

        When doc_divisors is provided, each item's amount is converted using
        its document's specific divisor BEFORE accumulation.  This correctly
        handles reports with mixed source units (e.g. FY2021=rupees, FY2022=lakhs).

        Returns a list of provenance records (one per item that was successfully processed).
        """
        provenance_records: list[dict] = []
        accumulator: dict[tuple[int, int], list[float]] = {}
        _converted_count = 0
        _raw_count = 0
        _skipped_count = 0

        for item in cell_data:
            field = item.get("cma_field_name")
            year = item.get("financial_year")
            amount = float(item.get("amount") or 0)

            # Prefer cma_row from classification (set by pipeline), fall back to static mapping
            row = item.get("cma_row") or ALL_FIELD_TO_ROW.get(field)
            col_letter = year_map.get(year)
            if not row or row == 0 or col_letter is None:
                if not row or row == 0:
                    logger.debug("No valid row for field '%s' (cma_row=%s) — skipping", field, item.get("cma_row"))
                if col_letter is None:
                    logger.warning("Year %s not in year mapping — skipping data cell", year)
                _skipped_count += 1
                continue

            # Per-document unit conversion: convert BEFORE accumulation
            doc_id = item.get("document_id")
            if doc_divisors and doc_id and doc_id in doc_divisors:
                divisor = doc_divisors[doc_id]
                if divisor and divisor != 1:
                    amount = round(amount / divisor, 2)
                    _converted_count += 1
                else:
                    _raw_count += 1
            elif unit_divisor and unit_divisor != 1:
                # Fallback to single divisor (backwards compat for tests)
                amount = round(amount / unit_divisor, 2)
                _converted_count += 1
            else:
                _raw_count += 1
                if doc_divisors and doc_id and doc_id not in doc_divisors:
                    # Safety net: item references a document not in doc_divisors
                    logger.warning(
                        "No divisor found for document_id=%s (field=%s, amount=%s) — writing raw amount",
                        doc_id, field, item.get("amount"),
                    )

            provenance_records.append({
                "cma_row": row,
                "cma_column": col_letter,
                "financial_year": year,
                "line_item_id": item.get("line_item_id"),
                "classification_id": item.get("classification_id"),
                "source_text": item.get("source_text"),
                "raw_amount": float(item.get("amount") or 0),
                "converted_amount": amount,
                "document_id": doc_id,
            })

            col = column_index_from_string(col_letter)
            key = (row, col)
            accumulator.setdefault(key, []).append(amount)

        logger.info(
            "Unit conversion summary: %d converted, %d raw (divisor=1), %d skipped (no row/col)",
            _converted_count, _raw_count, _skipped_count,
        )

        for (row, col), values in accumulator.items():
            cell = ws.cell(row=row, column=col)
            existing = cell.value
            if isinstance(existing, str) and existing.startswith("="):
                formula_body = existing[1:].strip()  # strip the leading '='
                if _REAL_FORMULA_RE.search(formula_body):
                    logger.warning(
                        "Skipping real formula cell at row=%d col=%d formula=%s",
                        row,
                        col,
                        existing,
                    )
                    continue
                # Placeholder formula (e.g. =0, =0.00) — safe to overwrite
                logger.debug(
                    "Overwriting placeholder formula '%s' at row=%d col=%d",
                    existing,
                    row,
                    col,
                )

            if len(values) == 1:
                cell.value = values[0]
            else:
                cell.value = _build_formula(values)
                logger.debug(
                    "Multi-value formula at row=%d col=%d: %s (%d items)",
                    row, col, cell.value, len(values),
                )

        return provenance_records

    # ── Private: I/O helpers ──────────────────────────────────────────────

    def _save_upload_cleanup(self, wb, report_id: str, user_id: str) -> str:
        """Save workbook to temp file, upload to Supabase Storage, cleanup."""
        if not _UUID_RE.match(report_id):
            raise ValueError(f"Invalid report_id format: {report_id!r}")
        storage_path = f"cma_reports/{report_id}/output.xlsm"
        tmp_path: str | None = None

        try:
            # mkstemp gives us a file descriptor + path; we close fd so openpyxl can write
            tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsm")
            os.close(tmp_fd)

            # Save — extension .xlsm preserves VBA/macro container
            wb.save(tmp_path)

            # Upload to Supabase Storage (bucket: generated)
            with open(tmp_path, "rb") as f:
                file_bytes = f.read()

            self.service.storage.from_(STORAGE_BUCKET).upload(
                path=storage_path,
                file=file_bytes,
                file_options={
                    "content-type": "application/vnd.ms-excel.sheet.macroenabled.12",
                    "upsert": "true",
                },
            )

            # Audit log
            log_action(
                self.service,
                user_id,
                "excel_generated",
                "cma_report",
                report_id,
                after={"output_path": storage_path},
            )

            return storage_path

        finally:
            if tmp_path:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    logger.warning("Could not delete temp file: %s", tmp_path)

    # ── Private: DB fetch helpers ─────────────────────────────────────────

    def _fetch_report(self, report_id: str) -> dict:
        result = (
            self.service.table("cma_reports")
            .select("id,client_id,document_ids,cma_output_unit")
            .eq("id", report_id)
            .single()
            .execute()
        )
        if not result.data:
            raise ValueError(f"CMA report not found: {report_id}")
        return result.data

    def _fetch_client_name(self, client_id: str) -> str:
        result = (
            self.service.table("clients")
            .select("id,name")
            .eq("id", client_id)
            .single()
            .execute()
        )
        if not result.data:
            raise ValueError(f"Client not found: {client_id}")
        return result.data["name"]

    def _fetch_documents(self, document_ids: list[str]) -> list[dict]:
        if not document_ids:
            return []
        result = (
            self.service.table("documents")
            .select("id,financial_year,nature,source_unit")
            .in_("id", document_ids)
            .execute()
        )
        return result.data or []

    def _fetch_classified_data(self, document_ids: list[str], doc_years: dict) -> list[dict]:
        """Return [{cma_field_name, financial_year, amount, document_id}] for all non-doubt classifications.

        Processes per-document to avoid PostgREST URL length limits
        (1000+ UUIDs in a single .in_() would exceed HTTP GET limits).
        Also paginates to bypass Supabase's default 1000-row limit.
        """
        if not document_ids or not doc_years:
            return []

        cell_data: list[dict] = []
        PAGE_SIZE = 1000

        for doc_id in document_ids:
            year = doc_years.get(doc_id)
            if not year:
                continue

            # Paginate line items for this document
            all_items: list[dict] = []
            offset = 0
            while True:
                page = (
                    self.service.table("extracted_line_items")
                    .select("id,document_id,amount")
                    .eq("document_id", doc_id)
                    .range(offset, offset + PAGE_SIZE - 1)
                    .execute()
                )
                batch = page.data or []
                all_items.extend(batch)
                if len(batch) < PAGE_SIZE:
                    break
                offset += PAGE_SIZE

            if not all_items:
                continue

            items = {row["id"]: row for row in all_items}
            item_ids = list(items.keys())

            # Batch classifications query (100 IDs per batch to stay within URL limits)
            _BATCH = 100
            classifications: list[dict] = []
            for i in range(0, len(item_ids), _BATCH):
                batch_ids = item_ids[i : i + _BATCH]
                clf_result = (
                    self.service.table("classifications")
                    .select("id,line_item_id,cma_field_name,cma_row")
                    .in_("line_item_id", batch_ids)
                    .eq("is_doubt", False)
                    .execute()
                )
                classifications.extend(clf_result.data or [])

            for clf in classifications:
                if not clf.get("cma_field_name"):
                    continue
                item = items.get(clf["line_item_id"])
                if not item:
                    continue
                cell_data.append(
                    {
                        "cma_field_name": clf["cma_field_name"],
                        "cma_row": clf.get("cma_row"),
                        "financial_year": year,
                        "amount": item.get("amount") or 0.0,
                        "document_id": doc_id,
                        "line_item_id": clf["line_item_id"],
                        "classification_id": clf["id"],
                        "source_text": item.get("description") or item.get("source_text"),
                    }
                )

            logger.info(
                "Fetched %d classified items for doc %s (FY%s)",
                len([c for c in classifications if c.get("cma_field_name")]),
                doc_id,
                year,
            )

        return cell_data
