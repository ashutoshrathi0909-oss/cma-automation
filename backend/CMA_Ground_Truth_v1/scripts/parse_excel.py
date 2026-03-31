#!/usr/bin/env python3
"""
Parse an Excel workbook into JSON that Claude Code subagents can read.

Usage:
    python parse_excel.py "Company File.xlsx" --output company_sheets.json
    python parse_excel.py "Company File.xlsx"  # prints to stdout

Requires: pip install openpyxl
"""

import argparse
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def parse_sheet(ws):
    """Extract all data from a worksheet, preserving merged cells."""
    rows = []
    merged_ranges = list(ws.merged_cells.ranges)

    # Build a map of merged cells to their top-left value
    merged_map = {}
    for mr in merged_ranges:
        top_left_value = ws.cell(mr.min_row, mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                if (row, col) != (mr.min_row, mr.min_col):
                    merged_map[(row, col)] = top_left_value

    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        row_data = []
        for cell in row:
            value = cell.value
            # If this cell is part of a merged range, use the merged value
            if (cell.row, cell.column) in merged_map:
                value = merged_map[(cell.row, cell.column)]

            # Convert to JSON-safe types
            if value is None:
                row_data.append(None)
            elif isinstance(value, (int, float)):
                row_data.append(value)
            else:
                row_data.append(str(value))
        rows.append(row_data)

    return rows


def parse_workbook(filepath):
    """Parse entire workbook into a dict of sheet_name -> rows."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=False)

    result = {
        "file": Path(filepath).name,
        "sheets": {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = parse_sheet(ws)

        # Skip completely empty sheets
        if not any(any(cell is not None for cell in row) for row in rows):
            continue

        # Trim trailing empty rows
        while rows and all(cell is None for cell in rows[-1]):
            rows.pop()

        result["sheets"][sheet_name] = {
            "row_count": len(rows),
            "col_count": ws.max_column,
            "data": rows
        }

    wb.close()
    return result


def main():
    parser = argparse.ArgumentParser(description="Parse Excel workbook to JSON")
    parser.add_argument("file", help="Path to .xlsx file")
    parser.add_argument("--output", "-o", help="Output JSON file (default: stdout)")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"Error: File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    result = parse_workbook(args.file)

    # Summary
    print(f"Parsed: {result['file']}", file=sys.stderr)
    for name, sheet in result["sheets"].items():
        print(f"  Sheet '{name}': {sheet['row_count']} rows x {sheet['col_count']} cols", file=sys.stderr)

    output = json.dumps(result, indent=2, ensure_ascii=False)

    if args.output:
        Path(args.output).write_text(output, encoding="utf-8")
        print(f"Saved to: {args.output}", file=sys.stderr)
    else:
        print(output)


if __name__ == "__main__":
    main()
