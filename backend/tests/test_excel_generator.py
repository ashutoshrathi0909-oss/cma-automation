"""Tests for ExcelGenerator — Phase 7.

Split into two groups:
  1. Pure layer tests (use real openpyxl workbook, no DB/file I/O mocking)
     — test fill_workbook(), _fill_headers(), _fill_data_cells()
  2. Integration layer tests (mock DB, file I/O, storage)
     — test generate(), _save_upload_cleanup()
"""

from __future__ import annotations

import os
import tempfile
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from app.services.excel_generator import (
    ExcelGenerator,
    _build_formula,
    _format_number,
    compute_unit_divisor,
)


# ── Helpers ────────────────────────────────────────────────────────────────


def _make_ws() -> tuple[openpyxl.Workbook, object]:
    """Return a minimal workbook + its INPUT SHEET worksheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INPUT SHEET"
    return wb, ws


def _make_generator() -> ExcelGenerator:
    """ExcelGenerator with mocked service (pure layer tests don't need DB)."""
    return ExcelGenerator(service=MagicMock(), template_path="/fake/template.xlsm")


# ── Pure layer tests ───────────────────────────────────────────────────────


def test_generator_fills_client_name_in_header():
    """Client name is written to row 7, column 1 (label column A)."""
    _, ws = _make_ws()
    gen = _make_generator()

    gen.fill_workbook(ws, client_name="Sharma & Sons Ltd", docs=[], cell_data=[])

    assert ws.cell(row=7, column=1).value == "Sharma & Sons Ltd"


def test_generator_fills_year_headers():
    """All 7 year columns (B-H) rewritten starting from base year."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=[])

    # base=2024 → B=2024, C=2025, D=2026, ..., H=2030
    assert ws.cell(row=8, column=2).value == 2024
    assert ws.cell(row=8, column=3).value == 2025
    assert ws.cell(row=8, column=8).value == 2030
    # Nature: 2024 is historical, rest are projected
    assert ws.cell(row=10, column=2).value == "audited"
    assert ws.cell(row=10, column=3).value == "Projected"


def test_generator_fills_nature_of_financials():
    """Historical docs get their nature; projection columns get 'Projected'."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [
        {"financial_year": 2024, "nature": "audited"},
        {"financial_year": 2025, "nature": "provisional"},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=[])

    # base=2024 → B=2024(audited), C=2025(provisional), D+=Projected
    assert ws.cell(row=10, column=2).value == "audited"
    assert ws.cell(row=10, column=3).value == "provisional"
    assert ws.cell(row=10, column=4).value == "Projected"


def test_generator_fills_pnl_domestic_sales_row_22():
    """'Domestic Sales' (PNL row 22) is filled for financial year 2024."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [{"cma_field_name": "Domestic Sales", "financial_year": 2024, "amount": 500_000.0}]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Only year is 2024 → base=2024 → column B → index 2
    assert ws.cell(row=22, column=2).value == 500_000.0


def test_generator_fills_pnl_wages_row_45():
    """'Wages' (PNL row 45) is filled correctly."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2023, "nature": "audited"}]
    cell_data = [{"cma_field_name": "Wages", "financial_year": 2023, "amount": 120_000.0}]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Only year is 2023 → base=2023 → column B → index 2
    assert ws.cell(row=45, column=2).value == 120_000.0


def test_generator_fills_bs_share_capital_row_116():
    """'Issued, Subscribed and Paid up' (BS row 116) is filled correctly."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {
            "cma_field_name": "Issued, Subscribed and Paid up",
            "financial_year": 2024,
            "amount": 1_000_000.0,
        }
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Only year is 2024 → base=2024 → column B → index 2
    assert ws.cell(row=116, column=2).value == 1_000_000.0


def test_generator_fills_bs_debtors_row_206():
    """'Domestic Receivables' (BS row 206) is filled correctly."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2025, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Domestic Receivables", "financial_year": 2025, "amount": 250_000.0}
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Only year is 2025 → base=2025 → column B → index 2
    assert ws.cell(row=206, column=2).value == 250_000.0


def test_generator_maps_year_to_correct_column():
    """Each year maps to the correct Excel column dynamically."""
    _, ws = _make_ws()
    gen = _make_generator()

    # Docs with 4 years → base=2023 → 2023=B(2), 2024=C(3), 2025=D(4), 2026=E(5)
    years = [2023, 2024, 2025, 2026]
    docs = [{"financial_year": yr, "nature": "audited"} for yr in years]
    year_to_expected_col = {2023: 2, 2024: 3, 2025: 4, 2026: 5}
    cell_data = [
        {"cma_field_name": "Domestic Sales", "financial_year": yr, "amount": 1.0}
        for yr in years
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    for yr, col in year_to_expected_col.items():
        assert ws.cell(row=22, column=col).value == 1.0, f"Year {yr} → col {col} failed"


def test_generator_handles_multiple_years():
    """Data for different years lands in different columns."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [
        {"financial_year": 2024, "nature": "audited"},
        {"financial_year": 2025, "nature": "provisional"},
    ]
    cell_data = [
        {"cma_field_name": "Wages", "financial_year": 2024, "amount": 100.0},
        {"cma_field_name": "Wages", "financial_year": 2025, "amount": 200.0},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # base=2024 → 2024=B(2), 2025=C(3)
    assert ws.cell(row=45, column=2).value == 100.0  # 2024 → col B
    assert ws.cell(row=45, column=3).value == 200.0  # 2025 → col C


def test_generator_writes_formula_for_multiple_items_same_row():
    """Multiple line items mapping to the same (field_name, year) produce an Excel formula."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Domestic Sales", "financial_year": 2024, "amount": 300_000.0},
        {"cma_field_name": "Domestic Sales", "financial_year": 2024, "amount": 200_000.0},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Instead of writing the sum 500000, writes a formula showing individual components
    assert ws.cell(row=22, column=2).value == "=300000+200000"


def test_generator_formula_handles_negative_amounts():
    """Negative amounts appear as subtraction in the formula."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Others (Admin)", "financial_year": 2024, "amount": 50_000.0},
        {"cma_field_name": "Others (Admin)", "financial_year": 2024, "amount": 30_000.0},
        {"cma_field_name": "Others (Admin)", "financial_year": 2024, "amount": -5_000.0},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # row 71 = Others (Admin), base=2024 → col B (index 2)
    assert ws.cell(row=71, column=2).value == "=50000+30000-5000"


def test_generator_single_item_writes_direct_value():
    """A single line item is written as a plain number, not a formula."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Domestic Sales", "financial_year": 2024, "amount": 500_000.0},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    assert ws.cell(row=22, column=2).value == 500_000.0


def test_generator_formula_preserves_decimals():
    """Decimal amounts are shown with 2 decimal places in the formula."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Wages", "financial_year": 2024, "amount": 100.50},
        {"cma_field_name": "Wages", "financial_year": 2024, "amount": 200.75},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    assert ws.cell(row=45, column=2).value == "=100.50+200.75"


def test_generator_does_not_overwrite_formula_cells():
    """Cells not in PNL_FIELD_TO_ROW / BS_FIELD_TO_ROW are left untouched."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2023, "nature": "audited"}]

    # Row 21 is NOT in PNL_FIELD_TO_ROW (first mapping entry is row 22)
    ws.cell(row=21, column=2).value = "=SUM(B17:B20)"
    cell_data = [
        {"cma_field_name": "Domestic Sales", "financial_year": 2023, "amount": 999.0}
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Only year is 2023 → base=2023 → column B → index 2
    assert ws.cell(row=22, column=2).value == 999.0
    # Row 21 formula should be UNTOUCHED
    assert ws.cell(row=21, column=2).value == "=SUM(B17:B20)"


# ── Integration layer tests ────────────────────────────────────────────────
# These mock: openpyxl.load_workbook, service.storage, os.unlink, log_action


def _build_full_mock_service() -> MagicMock:
    """Builds a mock Supabase service with all needed chains pre-configured."""
    service = MagicMock()

    report_row = {
        "id": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
        "client_id": "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb",
        "document_ids": ["doc-ccc"],
    }
    client_row = {"id": "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb", "name": "Test Client Ltd"}
    doc_row = {"id": "doc-ccc", "financial_year": 2024, "nature": "audited"}
    item_row = {"id": "item-ddd", "document_id": "doc-ccc", "amount": 100.0}
    clf_row = {"line_item_id": "item-ddd", "cma_field_name": "Domestic Sales"}

    def table_side_effect(name: str):
        chain = MagicMock()
        chain.select.return_value = chain
        chain.eq.return_value = chain
        chain.in_.return_value = chain
        chain.single.return_value = chain
        chain.update.return_value = chain

        if name == "cma_reports":
            chain.execute.return_value = MagicMock(data=report_row)
        elif name == "clients":
            chain.execute.return_value = MagicMock(data=client_row)
        elif name == "documents":
            chain.execute.return_value = MagicMock(data=[doc_row])
        elif name == "extracted_line_items":
            chain.execute.return_value = MagicMock(data=[item_row])
        elif name == "classifications":
            chain.execute.return_value = MagicMock(data=[clf_row])
        else:
            chain.execute.return_value = MagicMock(data=[])
        return chain

    service.table.side_effect = table_side_effect

    storage_bucket = MagicMock()
    storage_bucket.upload.return_value = {"Key": "cma_reports/aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa/output.xlsm"}
    service.storage.from_.return_value = storage_bucket

    return service


def _make_mock_wb():
    """Minimal mock workbook + worksheet for integration tests."""
    ws = MagicMock()
    cell_mock = MagicMock()
    ws.cell.return_value = cell_mock
    wb = MagicMock()
    wb.__getitem__ = lambda s, k: ws
    return wb, ws


def test_generator_opens_template_with_keep_vba():
    """load_workbook is called with keep_vba=True."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        mock_load.return_value, _ = _make_mock_wb()
        gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
        gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    mock_load.assert_called_once_with("/fake/CMA.xlsm", keep_vba=True)


def test_generator_preserves_macros():
    """Same as above: keep_vba=True ensures VBA/macros are retained."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        mock_load.return_value, _ = _make_mock_wb()
        gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
        gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    _, kwargs = mock_load.call_args
    assert kwargs.get("keep_vba") is True


def test_generator_saves_as_xlsm():
    """Workbook is saved with a .xlsm extension (never .xlsx)."""
    service = _build_full_mock_service()
    saved_paths = []

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        wb_mock, _ = _make_mock_wb()
        wb_mock.save.side_effect = lambda p: saved_paths.append(p)
        mock_load.return_value = wb_mock

        gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
        gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    assert saved_paths, "wb.save was never called"
    assert saved_paths[0].endswith(".xlsm"), f"Expected .xlsm, got: {saved_paths[0]}"


def test_generator_uploads_to_supabase_storage():
    """Generated file is uploaded to the 'generated' Supabase Storage bucket."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        wb_mock, _ = _make_mock_wb()
        mock_load.return_value = wb_mock

        gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
        storage_path = gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    service.storage.from_.assert_called_with("generated")
    service.storage.from_().upload.assert_called_once()
    assert storage_path == "cma_reports/aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa/output.xlsm"


def test_generator_logs_audit_trail():
    """log_action is called after successful upload."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        wb_mock, _ = _make_mock_wb()
        mock_load.return_value = wb_mock

        with patch("app.services.excel_generator.log_action") as mock_log:
            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    mock_log.assert_called_once()
    args = mock_log.call_args[0]
    assert args[2] == "excel_generated"
    assert args[3] == "cma_report"
    assert args[4] == "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"


def test_generator_cleans_up_temp_file():
    """Temp .xlsm file is deleted after upload, even on success."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        wb_mock, _ = _make_mock_wb()
        mock_load.return_value = wb_mock

        with patch("app.services.excel_generator.os.unlink") as mock_unlink:
            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    mock_unlink.assert_called_once()
    unlinked = mock_unlink.call_args[0][0]
    assert unlinked.endswith(".xlsm"), f"Expected .xlsm cleanup, got: {unlinked}"


# ── Branch coverage tests (edge cases for 100% coverage) ──────────────────


def test_fill_headers_writes_all_years_from_any_base():
    """Any base year maps correctly — B through H get sequential years."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 1900, "nature": "audited"}]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=[])

    # base=1900 → B=1900, C=1901, ..., H=1906
    assert ws.cell(row=8, column=2).value == 1900
    assert ws.cell(row=8, column=8).value == 1906
    assert ws.cell(row=10, column=2).value == "audited"
    assert ws.cell(row=10, column=3).value == "Projected"


def test_fill_data_cells_skips_unknown_field_and_year():
    """_fill_data_cells silently skips items with unknown field or year."""
    _, ws = _make_ws()
    gen = _make_generator()
    docs = [{"financial_year": 2024, "nature": "audited"}]
    cell_data = [
        {"cma_field_name": "Nonexistent Field XYZ", "financial_year": 2024, "amount": 999.0},
        {"cma_field_name": "Domestic Sales", "financial_year": 1900, "amount": 888.0},
    ]

    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Unknown field: skipped. Year 1900 not in year_map (only 2024): skipped.
    # Row 22 col B (index 2) should be None because the only valid-year item had unknown field
    for col in range(1, 10):
        assert ws.cell(row=22, column=col).value is None


def test_fetch_report_raises_when_not_found():
    """_fetch_report raises ValueError when the report row is missing."""
    service = MagicMock()
    chain = MagicMock()
    chain.select.return_value = chain
    chain.eq.return_value = chain
    chain.single.return_value = chain
    chain.execute.return_value = MagicMock(data=None)
    service.table.return_value = chain

    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    with pytest.raises(ValueError, match="CMA report not found"):
        gen._fetch_report("missing-id")


def test_fetch_client_name_raises_when_not_found():
    """_fetch_client_name raises ValueError when the client row is missing."""
    service = MagicMock()
    chain = MagicMock()
    chain.select.return_value = chain
    chain.eq.return_value = chain
    chain.single.return_value = chain
    chain.execute.return_value = MagicMock(data=None)
    service.table.return_value = chain

    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    with pytest.raises(ValueError, match="Client not found"):
        gen._fetch_client_name("missing-client")


def test_fetch_documents_returns_empty_for_no_ids():
    """_fetch_documents returns [] immediately when given an empty list."""
    gen = ExcelGenerator(service=MagicMock(), template_path="/fake/CMA.xlsm")
    result = gen._fetch_documents([])
    assert result == []


def test_fetch_classified_data_returns_empty_for_no_ids():
    """_fetch_classified_data returns [] immediately when given an empty list."""
    gen = ExcelGenerator(service=MagicMock(), template_path="/fake/CMA.xlsm")
    result = gen._fetch_classified_data([], {})
    assert result == []


def test_fetch_classified_data_returns_empty_when_no_items():
    """_fetch_classified_data returns [] when no extracted_line_items exist."""
    service = MagicMock()
    chain = MagicMock()
    chain.select.return_value = chain
    chain.in_.return_value = chain
    chain.execute.return_value = MagicMock(data=[])
    service.table.return_value = chain

    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    result = gen._fetch_classified_data(["doc-1"], {"doc-1": 2024})
    assert result == []


def test_fetch_classified_data_skips_clf_without_field_name():
    """_fetch_classified_data skips classifications with no cma_field_name."""
    service = MagicMock()
    item_row = {"id": "item-1", "document_id": "doc-1", "amount": 50.0}
    clf_row_no_field = {"line_item_id": "item-1", "cma_field_name": None}

    def table_side_effect(name):
        chain = MagicMock()
        chain.select.return_value = chain
        chain.in_.return_value = chain
        chain.eq.return_value = chain
        chain.execute.return_value = MagicMock(data=[])
        if name == "extracted_line_items":
            chain.execute.return_value = MagicMock(data=[item_row])
        elif name == "classifications":
            chain.execute.return_value = MagicMock(data=[clf_row_no_field])
        return chain

    service.table.side_effect = table_side_effect
    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    result = gen._fetch_classified_data(["doc-1"], {"doc-1": 2024})
    assert result == []


def test_fetch_classified_data_skips_clf_with_unknown_item():
    """_fetch_classified_data skips classifications whose line_item_id is not in items."""
    service = MagicMock()
    item_row = {"id": "item-1", "document_id": "doc-1", "amount": 50.0}
    clf_row_orphan = {"line_item_id": "item-UNKNOWN", "cma_field_name": "Wages"}

    def table_side_effect(name):
        chain = MagicMock()
        chain.select.return_value = chain
        chain.in_.return_value = chain
        chain.eq.return_value = chain
        chain.execute.return_value = MagicMock(data=[])
        if name == "extracted_line_items":
            chain.execute.return_value = MagicMock(data=[item_row])
        elif name == "classifications":
            chain.execute.return_value = MagicMock(data=[clf_row_orphan])
        return chain

    service.table.side_effect = table_side_effect
    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    result = gen._fetch_classified_data(["doc-1"], {"doc-1": 2024})
    assert result == []


def test_fetch_classified_data_skips_item_without_year():
    """_fetch_classified_data skips items whose document has no year mapping."""
    service = MagicMock()
    item_row = {"id": "item-1", "document_id": "doc-ORPHAN", "amount": 50.0}
    clf_row = {"line_item_id": "item-1", "cma_field_name": "Wages"}
    # doc_years does not contain doc-ORPHAN → year lookup returns None → item skipped

    def table_side_effect(name):
        chain = MagicMock()
        chain.select.return_value = chain
        chain.in_.return_value = chain
        chain.eq.return_value = chain
        chain.execute.return_value = MagicMock(data=[])
        if name == "extracted_line_items":
            chain.execute.return_value = MagicMock(data=[item_row])
        elif name == "classifications":
            chain.execute.return_value = MagicMock(data=[clf_row])
        return chain

    service.table.side_effect = table_side_effect
    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    # Pass a non-empty doc_years but with a different doc id — year lookup will miss
    result = gen._fetch_classified_data(["doc-ORPHAN"], {"doc-OTHER": 2024})
    assert result == []


def test_fetch_classified_data_filters_out_doubt_items():
    """_fetch_classified_data calls .eq('is_doubt', False) on classifications query — doubt items never silently classified."""
    service = MagicMock()

    # Items chain — must support .select().eq().range().execute() for paginated fetch
    items_chain = MagicMock()
    items_chain.select.return_value = items_chain
    items_chain.eq.return_value = items_chain
    items_chain.in_.return_value = items_chain
    items_chain.range.return_value = items_chain
    items_chain.execute.return_value = MagicMock(data=[{"id": "item-1", "document_id": "doc-1", "amount": 100.0}])

    # Classifications chain — capture the eq call (.select().in_().eq().execute())
    clf_chain = MagicMock()
    clf_chain.select.return_value = clf_chain
    clf_chain.in_.return_value = clf_chain
    clf_chain.eq.return_value = clf_chain
    clf_chain.execute.return_value = MagicMock(data=[])

    def table_se(name):
        if name == "extracted_line_items":
            return items_chain
        elif name == "classifications":
            return clf_chain
        chain = MagicMock()
        chain.select.return_value = chain
        chain.in_.return_value = chain
        chain.execute.return_value = MagicMock(data=[])
        return chain

    service.table.side_effect = table_se

    gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
    doc_years = {"doc-1": 2024}
    gen._fetch_classified_data(["doc-1"], doc_years)

    # CRITICAL: verify is_doubt=False filter was applied
    clf_chain.eq.assert_called_once_with("is_doubt", False)


def test_generator_swallows_oserror_on_temp_cleanup():
    """OSError during temp file cleanup is swallowed (logged as warning, not raised)."""
    service = _build_full_mock_service()

    with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
        wb_mock, _ = _make_mock_wb()
        mock_load.return_value = wb_mock

        with patch("app.services.excel_generator.os.unlink", side_effect=OSError("locked")):
            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            # Should NOT raise — OSError is caught and logged
            storage_path = gen.generate(report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa", user_id="user-xxx")

    assert storage_path == "cma_reports/aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa/output.xlsm"


def test_save_upload_cleanup_rejects_non_uuid_report_id():
    """_save_upload_cleanup raises ValueError for non-UUID report_id (path traversal guard)."""
    gen = ExcelGenerator(service=MagicMock(), template_path="/fake/CMA.xlsm")
    wb = openpyxl.Workbook()

    with pytest.raises(ValueError, match="Invalid report_id format"):
        gen._save_upload_cleanup(wb, "../../etc/passwd", "user-xxx")


def test_fill_data_cells_skips_formula_cells():
    """_fill_data_cells does not overwrite cells that already contain a real formula."""
    _, ws = _make_ws()
    # Pre-set a formula in the Domestic Sales cell (row 22, col B = col 2, for year 2023)
    ws.cell(row=22, column=2).value = "=SUM(C22:D22)"
    gen = _make_generator()
    docs = [{"financial_year": 2023, "nature": "audited"}]

    cell_data = [{"cma_field_name": "Domestic Sales", "financial_year": 2023, "amount": 99999.0}]
    gen.fill_workbook(ws, client_name="Test Co", docs=docs, cell_data=cell_data)

    # Real formula must be preserved — not overwritten with the data amount
    assert ws.cell(row=22, column=2).value == "=SUM(C22:D22)"


# ── Helper function unit tests ────────────────────────────────────────────


def test_format_number_integer():
    assert _format_number(5000.0) == "5000"


def test_format_number_decimal():
    assert _format_number(100.50) == "100.50"


def test_format_number_negative():
    assert _format_number(-3000.0) == "-3000"


def test_build_formula_single_value():
    assert _build_formula([50000]) == "=50000"


def test_build_formula_multiple_positive():
    assert _build_formula([300000, 200000]) == "=300000+200000"


def test_build_formula_with_negatives():
    assert _build_formula([50000, 30000, -5000]) == "=50000+30000-5000"


def test_build_formula_all_negative():
    assert _build_formula([-1000, -2000]) == "=-1000-2000"


def test_build_formula_with_decimals():
    assert _build_formula([100.50, 200.75]) == "=100.50+200.75"


def test_build_formula_empty_list():
    assert _build_formula([]) == "=0"


def test_build_formula_first_negative_second_positive():
    assert _build_formula([-5000, 3000]) == "=-5000+3000"


def test_build_formula_zeros():
    assert _build_formula([0, 5000]) == "=0+5000"


def test_format_number_nan():
    assert _format_number(float("nan")) == "0"


def test_format_number_inf():
    assert _format_number(float("inf")) == "0"


# ── Unit conversion tests ────────────────────────────────────────────────


class TestComputeUnitDivisor:
    """Tests for compute_unit_divisor — the heart of per-document conversion."""

    def test_lakhs_to_crores(self):
        assert compute_unit_divisor("lakhs", "crores") == 100

    def test_rupees_to_crores(self):
        assert compute_unit_divisor("rupees", "crores") == 10_000_000

    def test_rupees_to_lakhs(self):
        assert compute_unit_divisor("rupees", "lakhs") == 100_000

    def test_thousands_to_lakhs(self):
        assert compute_unit_divisor("thousands", "lakhs") == 100

    def test_thousands_to_crores(self):
        assert compute_unit_divisor("thousands", "crores") == 10_000

    def test_lakhs_to_lakhs(self):
        """Same unit → divisor 1 → no conversion."""
        assert compute_unit_divisor("lakhs", "lakhs") == 1

    def test_crores_to_crores(self):
        """Same unit → divisor 1 → no conversion."""
        assert compute_unit_divisor("crores", "crores") == 1

    def test_crores_to_lakhs(self):
        """Crores→lakhs means MULTIPLYING, so divisor < 1."""
        assert compute_unit_divisor("crores", "lakhs") == pytest.approx(0.01)

    def test_unknown_source_defaults_to_1(self):
        """Unknown source_unit defaults to 1 (rupees)."""
        assert compute_unit_divisor("unknown_unit", "crores") == 10_000_000

    def test_unknown_output_defaults_to_lakhs(self):
        """Unknown output_unit defaults to lakhs (100_000)."""
        assert compute_unit_divisor("rupees", "unknown_unit") == 100_000


class TestFillDataCellsWithDocDivisors:
    """Tests for _fill_data_cells unit conversion using per-document divisors.

    This is the critical path: each item's amount must be divided by its
    document's divisor BEFORE writing to the worksheet.
    """

    def test_single_doc_lakhs_to_crores(self):
        """Lakhs source → crores output: amounts divided by 100."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 15000.0,  # 15000 lakhs
                "document_id": "doc-fy22",
            }
        ]
        doc_divisors = {"doc-fy22": 100}  # lakhs → crores

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # 15000 / 100 = 150.0 crores
        assert ws.cell(row=22, column=2).value == 150.0

    def test_single_doc_rupees_to_crores(self):
        """Rupees source → crores output: amounts divided by 10,000,000."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2023, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Wages",
                "financial_year": 2023,
                "amount": 50_000_000.0,  # 5 crore in rupees
                "document_id": "doc-fy23",
            }
        ]
        doc_divisors = {"doc-fy23": 10_000_000}  # rupees → crores

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # 50,000,000 / 10,000,000 = 5.0 crores
        assert ws.cell(row=45, column=2).value == 5.0

    def test_mixed_source_units_across_years(self):
        """FY2021 in rupees, FY2022 in lakhs, output in crores.

        This is the exact scenario from the bug report:
        each year has a different source_unit, and the generator must
        apply the correct divisor per-document.
        """
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [
            {"financial_year": 2021, "nature": "audited"},
            {"financial_year": 2022, "nature": "audited"},
            {"financial_year": 2023, "nature": "audited"},
        ]
        cell_data = [
            # FY2021: rupees
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2021,
                "amount": 100_000_000.0,  # 10 crore in rupees
                "document_id": "doc-fy21",
            },
            # FY2022: lakhs (the bug scenario)
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 23700.0,  # 237 crore in lakhs
                "document_id": "doc-fy22",
            },
            # FY2023: rupees
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2023,
                "amount": 150_000_000.0,  # 15 crore in rupees
                "document_id": "doc-fy23",
            },
        ]
        doc_divisors = {
            "doc-fy21": 10_000_000,  # rupees → crores
            "doc-fy22": 100,          # lakhs → crores
            "doc-fy23": 10_000_000,  # rupees → crores
        }

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # base=2021 → B=2021, C=2022, D=2023
        assert ws.cell(row=22, column=2).value == 10.0   # 100M / 10M = 10 Cr
        assert ws.cell(row=22, column=3).value == 237.0   # 23700 / 100 = 237 Cr
        assert ws.cell(row=22, column=4).value == 15.0    # 150M / 10M = 15 Cr

    def test_no_conversion_when_divisor_is_1(self):
        """When divisor is 1 (same unit), amounts pass through unchanged."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2024, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2024,
                "amount": 500.0,
                "document_id": "doc-1",
            }
        ]
        doc_divisors = {"doc-1": 1}  # lakhs → lakhs = no conversion

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        assert ws.cell(row=22, column=2).value == 500.0

    def test_multi_item_formula_with_conversion(self):
        """Multiple items mapping to same cell: each value converted before formula."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Salary and staff expenses",
                "financial_year": 2022,
                "amount": 15471.0,  # raw lakhs
                "document_id": "doc-fy22",
            },
            {
                "cma_field_name": "Salary and staff expenses",
                "financial_year": 2022,
                "amount": 11.98,  # raw lakhs
                "document_id": "doc-fy22",
            },
        ]
        doc_divisors = {"doc-fy22": 100}  # lakhs → crores

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # 15471/100 = 154.71, 11.98/100 = 0.12
        # row 67 = Salary and staff expenses, col B (base=2022)
        assert ws.cell(row=67, column=2).value == "=154.71+0.12"

    def test_conversion_applied_to_all_sections(self):
        """Conversion works for both P&L and Balance Sheet items."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            # P&L item
            {
                "cma_field_name": "Wages",
                "financial_year": 2022,
                "amount": 500.0,  # lakhs
                "document_id": "doc-1",
            },
            # Balance Sheet item
            {
                "cma_field_name": "Bank Balances",
                "financial_year": 2022,
                "amount": 11.28,  # lakhs
                "document_id": "doc-1",
            },
        ]
        doc_divisors = {"doc-1": 100}  # lakhs → crores

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # Wages (row 45): 500/100 = 5.0
        assert ws.cell(row=45, column=2).value == 5.0
        # Bank Balances (row 213): 11.28/100 = 0.11 (rounded to 2dp)
        assert ws.cell(row=213, column=2).value == 0.11

    def test_fallback_to_single_unit_divisor(self):
        """When doc_divisors is None, falls back to the legacy single unit_divisor."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 10000.0,
                "document_id": "doc-1",
            }
        ]

        gen.fill_workbook(ws, "Test Co", docs, cell_data, unit_divisor=100)

        # 10000 / 100 = 100.0
        assert ws.cell(row=22, column=2).value == 100.0

    def test_item_without_document_id_uses_fallback(self):
        """Item missing document_id falls back to single unit_divisor."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 10000.0,
                # no document_id
            }
        ]
        doc_divisors = {"doc-1": 100}

        gen.fill_workbook(ws, "Test Co", docs, cell_data, unit_divisor=50, doc_divisors=doc_divisors)

        # No matching doc_id in doc_divisors → falls back to unit_divisor=50
        # 10000 / 50 = 200.0
        assert ws.cell(row=22, column=2).value == 200.0

    def test_item_with_unknown_document_id_uses_fallback(self):
        """Item with doc_id not in doc_divisors falls back to single unit_divisor."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 10000.0,
                "document_id": "doc-unknown",
            }
        ]
        doc_divisors = {"doc-1": 100}

        gen.fill_workbook(ws, "Test Co", docs, cell_data, unit_divisor=50, doc_divisors=doc_divisors)

        # doc-unknown not in doc_divisors → falls back to unit_divisor=50
        # 10000 / 50 = 200.0
        assert ws.cell(row=22, column=2).value == 200.0

    def test_no_value_over_1000_for_lakhs_to_crores(self):
        """Bug regression: for BCIPL FY2022 in lakhs → crores, no cell should exceed ~237 Cr.

        If any value is > 1000, it's still in raw lakhs (not converted).
        """
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        # Simulate multiple items from a lakhs-denominated document
        cell_data = [
            {"cma_field_name": "Domestic Sales", "financial_year": 2022, "amount": 23700.0, "document_id": "doc-fy22"},
            {"cma_field_name": "Wages", "financial_year": 2022, "amount": 166.0, "document_id": "doc-fy22"},
            {"cma_field_name": "Bank Balances", "financial_year": 2022, "amount": 2.0, "document_id": "doc-fy22"},
            {"cma_field_name": "Audit Fees & Directors Remuneration", "financial_year": 2022, "amount": 15.0, "document_id": "doc-fy22"},
        ]
        doc_divisors = {"doc-fy22": 100}  # lakhs → crores

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors=doc_divisors)

        # Verify converted values
        assert ws.cell(row=22, column=2).value == 237.0    # 23700/100
        assert ws.cell(row=45, column=2).value == 1.66     # 166/100
        assert ws.cell(row=213, column=2).value == 0.02    # 2/100
        assert ws.cell(row=73, column=2).value == 0.15     # 15/100

    def test_item_without_doc_id_and_no_fallback_passes_raw(self):
        """Item without doc_id and unit_divisor=1: amount passes through raw."""
        _, ws = _make_ws()
        gen = _make_generator()
        docs = [{"financial_year": 2022, "nature": "audited"}]
        cell_data = [
            {
                "cma_field_name": "Domestic Sales",
                "financial_year": 2022,
                "amount": 15471.0,
                # no document_id
            }
        ]

        gen.fill_workbook(ws, "Test Co", docs, cell_data, doc_divisors={})

        # Empty doc_divisors + no fallback → raw amount
        assert ws.cell(row=22, column=2).value == 15471.0


class TestGenerateIntegrationWithUnits:
    """Integration tests verifying that generate() correctly computes and applies doc_divisors."""

    def _build_service_with_units(
        self,
        source_unit: str = "lakhs",
        cma_output_unit: str = "crores",
    ) -> MagicMock:
        """Build mock service with source_unit and cma_output_unit set."""
        service = MagicMock()

        report_row = {
            "id": "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
            "client_id": "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb",
            "document_ids": ["doc-ccc"],
            "cma_output_unit": cma_output_unit,
        }
        client_row = {"id": "bbbbbbbb-bbbb-bbbb-bbbb-bbbbbbbbbbbb", "name": "Test Client Ltd"}
        doc_row = {
            "id": "doc-ccc",
            "financial_year": 2022,
            "nature": "audited",
            "source_unit": source_unit,
        }
        item_row = {"id": "item-ddd", "document_id": "doc-ccc", "amount": 15000.0}
        clf_row = {"line_item_id": "item-ddd", "cma_field_name": "Domestic Sales", "cma_row": 22}

        call_count = {"extracted_line_items": 0}

        def table_side_effect(name: str):
            chain = MagicMock()
            chain.select.return_value = chain
            chain.eq.return_value = chain
            chain.in_.return_value = chain
            chain.single.return_value = chain
            chain.update.return_value = chain
            chain.range.return_value = chain

            if name == "cma_reports":
                chain.execute.return_value = MagicMock(data=report_row)
            elif name == "clients":
                chain.execute.return_value = MagicMock(data=client_row)
            elif name == "documents":
                chain.execute.return_value = MagicMock(data=[doc_row])
            elif name == "extracted_line_items":
                # First call returns items, second call returns empty (pagination end)
                call_count["extracted_line_items"] += 1
                if call_count["extracted_line_items"] == 1:
                    chain.execute.return_value = MagicMock(data=[item_row])
                else:
                    chain.execute.return_value = MagicMock(data=[])
            elif name == "classifications":
                chain.execute.return_value = MagicMock(data=[clf_row])
            elif name == "cma_report_history":
                chain.execute.return_value = MagicMock(data=[])
            else:
                chain.execute.return_value = MagicMock(data=[])
            return chain

        service.table.side_effect = table_side_effect

        storage_bucket = MagicMock()
        storage_bucket.upload.return_value = {"Key": "cma_reports/aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa/output.xlsm"}
        service.storage.from_.return_value = storage_bucket

        return service

    def test_generate_applies_lakhs_to_crores_conversion(self):
        """End-to-end: generate() with lakhs source + crores output divides by 100."""
        service = self._build_service_with_units(source_unit="lakhs", cma_output_unit="crores")

        # Use a real workbook to verify actual cell values
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "INPUT SHEET"

        with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
            mock_load.return_value = wb
            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            gen.generate(
                report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                user_id="user-xxx",
            )

        # Domestic Sales = row 22, FY2022 = col B (index 2, since only year)
        cell_val = ws.cell(row=22, column=2).value
        # 15000 lakhs / 100 = 150.0 crores
        assert cell_val == 150.0, f"Expected 150.0 (crores), got {cell_val} — lakhs not converted!"

    def test_generate_no_conversion_when_same_unit(self):
        """generate() with lakhs source + lakhs output: divisor=1, no conversion."""
        service = self._build_service_with_units(source_unit="lakhs", cma_output_unit="lakhs")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "INPUT SHEET"

        with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
            mock_load.return_value = wb
            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            gen.generate(
                report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                user_id="user-xxx",
            )

        cell_val = ws.cell(row=22, column=2).value
        # 15000 lakhs / 1 = 15000.0 lakhs
        assert cell_val == 15000.0, f"Expected 15000.0 (same unit), got {cell_val}"

    def test_generate_rupees_to_crores_conversion(self):
        """generate() with rupees source + crores output: divides by 10,000,000."""
        service = self._build_service_with_units(source_unit="rupees", cma_output_unit="crores")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "INPUT SHEET"

        with patch("app.services.excel_generator.openpyxl.load_workbook") as mock_load:
            mock_load.return_value = wb

            # Override item amount for this test
            # Find and update the mock to return a rupees amount
            original_side_effect = service.table.side_effect
            call_count = {"extracted_line_items": 0}

            def modified_table(name):
                chain = original_side_effect(name)
                if name == "extracted_line_items":
                    call_count["extracted_line_items"] += 1
                    if call_count["extracted_line_items"] <= 1:
                        chain.execute.return_value = MagicMock(
                            data=[{"id": "item-ddd", "document_id": "doc-ccc", "amount": 50_000_000.0}]
                        )
                return chain

            service.table.side_effect = modified_table

            gen = ExcelGenerator(service=service, template_path="/fake/CMA.xlsm")
            gen.generate(
                report_id="aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa",
                user_id="user-xxx",
            )

        cell_val = ws.cell(row=22, column=2).value
        # 50,000,000 rupees / 10,000,000 = 5.0 crores
        assert cell_val == 5.0, f"Expected 5.0 (crores), got {cell_val}"
