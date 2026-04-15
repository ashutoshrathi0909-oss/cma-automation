"""Test that cell_note is rendered as an openpyxl Comment on the target cell."""
from __future__ import annotations
import shutil
from pathlib import Path
import openpyxl
import pytest

REPO_ROOT = Path(__file__).resolve().parents[3]


@pytest.fixture
def template_copy(tmp_path):
    src = REPO_ROOT / "DOCS" / "CMA.xlsm"
    dst = tmp_path / "test.xlsm"
    shutil.copy(src, dst)
    return dst


def test_write_data_cell_with_note_attaches_comment(template_copy):
    from app.services.excel_generator import write_data_cell_with_note
    wb = openpyxl.load_workbook(str(template_copy), keep_vba=True)
    ws = wb["INPUT SHEET"]
    write_data_cell_with_note(
        ws, row=131, col_letter="B",
        value=500000,
        note="Includes bill discounting balance of ₹50,000 per Note 22",
    )
    cell = ws["B131"]
    assert cell.value == 500000
    assert cell.comment is not None
    assert "bill discounting" in cell.comment.text


def test_write_data_cell_with_note_no_note_no_comment(template_copy):
    from app.services.excel_generator import write_data_cell_with_note
    wb = openpyxl.load_workbook(str(template_copy), keep_vba=True)
    ws = wb["INPUT SHEET"]
    write_data_cell_with_note(ws, row=22, col_letter="B", value=1_000_000, note=None)
    cell = ws["B22"]
    assert cell.value == 1_000_000
    assert cell.comment is None


def test_write_data_cell_with_note_empty_note_no_comment(template_copy):
    """Empty string note should not produce a comment."""
    from app.services.excel_generator import write_data_cell_with_note
    wb = openpyxl.load_workbook(str(template_copy), keep_vba=True)
    ws = wb["INPUT SHEET"]
    write_data_cell_with_note(ws, row=22, col_letter="B", value=42, note="")
    cell = ws["B22"]
    assert cell.value == 42
    assert cell.comment is None
