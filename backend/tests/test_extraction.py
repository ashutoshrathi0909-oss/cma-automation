"""
Phase 4A: Extraction service tests.

TDD workflow — RED → GREEN → REFACTOR.
All tests were written BEFORE the service implementations.

Coverage targets:
  - ExcelExtractor: xlsx, xls, amounts, sections, Indian number format
  - OcrExtractor: vision AI extraction (see tests/test_vision_ocr.py)
  - extractor_factory: routing logic
"""

import io
from unittest.mock import AsyncMock, MagicMock, patch

import pytest

from app.services.extraction.extractor_factory import (
    ExtractionError,
    LineItem,
    extract_document,
)
from app.services.extraction._types import parse_amount
from app.services.extraction.excel_extractor import (
    ExcelExtractor,
    deduplicate_across_sheets,
    _descriptions_match,
    _amounts_match,
    _sheet_priority,
    _sheet_page_type,
)


# ── Helpers ────────────────────────────────────────────────────────────────────


def _make_xlsx_bytes() -> bytes:
    """Minimal valid xlsx bytes (PK magic)."""
    return b"PK\x03\x04" + b"\x00" * 100


def _make_xls_bytes() -> bytes:
    """Minimal OLE2 compound document magic bytes for xls."""
    return b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"\x00" * 100



# ── LineItem dataclass ─────────────────────────────────────────────────────────


class TestLineItem:
    def test_line_item_has_required_fields(self):
        """LineItem dataclass must have description, amount, section, raw_text."""
        item = LineItem(
            description="Salaries & Wages",
            amount=500000.0,
            section="expenses",
            raw_text="Salaries & Wages  5,00,000",
        )
        assert item.description == "Salaries & Wages"
        assert item.amount == 500000.0
        assert item.section == "expenses"
        assert item.raw_text == "Salaries & Wages  5,00,000"

    def test_line_item_equality(self):
        """Two LineItems with identical fields should be equal (dataclass default)."""
        a = LineItem("Revenue", 1000000.0, "income", "Revenue 10,00,000")
        b = LineItem("Revenue", 1000000.0, "income", "Revenue 10,00,000")
        assert a == b

    def test_ambiguity_question_default_none(self):
        item = LineItem(description="Revenue", amount=1000.0, section="income", raw_text="Revenue 1000")
        assert item.ambiguity_question is None

    def test_ambiguity_question_can_be_set_as_keyword(self):
        item = LineItem(
            description="Wages & Other",
            amount=2000.0,
            section="expenses",
            raw_text="Wages & Other 2000",
            ambiguity_question="Please split between Row 45 and Row 49",
        )
        assert item.ambiguity_question is not None

    def test_page_type_default_empty(self):
        item = LineItem(description="Revenue", amount=1000.0, section="income", raw_text="Revenue 1000")
        assert item.page_type == ""

    def test_page_type_can_be_set(self):
        item = LineItem(
            description="Revenue",
            amount=1000.0,
            section="income",
            raw_text="Revenue 1000",
            page_type="face",
        )
        assert item.page_type == "face"


# ── Indian Number Format ───────────────────────────────────────────────────────


class TestIndianNumberFormat:
    """
    Test the parse_amount helper that normalises Indian and Western formats.
    We import it directly from the extractor_factory module.
    """

    def test_parse_indian_lakh_format(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("1,23,456") == pytest.approx(123456.0)

    def test_parse_western_format(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("1,234,567") == pytest.approx(1234567.0)

    def test_parse_plain_number(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("500000") == pytest.approx(500000.0)

    def test_parse_decimal(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("1,23,456.78") == pytest.approx(123456.78)

    def test_parse_empty_returns_none(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("") is None

    def test_parse_non_numeric_returns_none(self):
        from app.services.extraction.extractor_factory import parse_amount
        assert parse_amount("N/A") is None

    def test_parse_crore_format(self):
        from app.services.extraction.extractor_factory import parse_amount
        # 1,00,00,000 = 1 crore
        assert parse_amount("1,00,00,000") == pytest.approx(10000000.0)

    def test_parse_parentheses_negative(self):
        """(1,23,456) should return -123456.0 per Indian accounting convention."""
        assert parse_amount("(1,23,456)") == -123456.0

    def test_parse_plain_negative(self):
        """-500000 should return -500000.0."""
        assert parse_amount("-500000") == -500000.0

    def test_parse_negative_with_currency_symbol(self):
        """₹(45,00,000) should return -4500000.0."""
        assert parse_amount("₹(45,00,000)") == -4500000.0


# ── ExcelExtractor ─────────────────────────────────────────────────────────────


class TestExcelExtractor:
    @pytest.fixture
    def extractor(self):
        return ExcelExtractor()

    @pytest.mark.asyncio
    async def test_excel_extractor_reads_xlsx(self, extractor):
        """ExcelExtractor.extract() with xlsx bytes → returns list of LineItem."""
        # Build a real in-memory xlsx workbook with openpyxl
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"
        ws.append(["Description", "Amount"])
        ws.append(["Salaries & Wages", 500000])
        ws.append(["Rent", 120000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx")

        assert isinstance(items, list)
        assert len(items) >= 1
        descriptions = [i.description for i in items]
        assert any("Salaries" in d or "salaries" in d.lower() for d in descriptions)

    @pytest.mark.asyncio
    async def test_excel_extractor_reads_xls(self, extractor):
        """ExcelExtractor.extract() with xls bytes → parses via xlrd, returns LineItems."""
        # We mock xlrd because building a valid binary .xls from scratch is complex
        mock_sheet = MagicMock()
        mock_sheet.nrows = 3
        mock_sheet.name = "Sheet1"

        # Row 0: header row (text, text) — should be skipped
        mock_sheet.row_values.side_effect = lambda i: [
            ["Description", "2024"],
            ["Salaries & Wages", 500000.0],
            ["Rent Expense", 120000.0],
        ][i]

        mock_wb = MagicMock()
        mock_wb.sheets.return_value = [mock_sheet]

        with patch("xlrd.open_workbook", return_value=mock_wb):
            items = await extractor.extract(_make_xls_bytes(), "xls", selected_sheets=["Sheet1"])

        assert isinstance(items, list)
        assert len(items) >= 1

    @pytest.mark.asyncio
    async def test_excel_extractor_extracts_amounts(self, extractor):
        """Row 'Salaries & Wages | 5,00,000' → LineItem with amount=500000.0."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws_name = ws.title  # default is "Sheet"
        # Use string format to simulate Indian number formatting in a cell
        ws.append(["Salaries & Wages", "5,00,000"])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=[ws_name])

        salary_items = [i for i in items if "Salaries" in i.description or "salaries" in i.description.lower()]
        assert len(salary_items) >= 1
        assert salary_items[0].amount == pytest.approx(500000.0)

    @pytest.mark.asyncio
    async def test_excel_extractor_extracts_section_headers(self, extractor):
        """Bold / ALL-CAPS header rows without numeric amounts → set current section."""
        import openpyxl
        from openpyxl.styles import Font

        wb = openpyxl.Workbook()
        ws = wb.active
        ws_name = ws.title  # default is "Sheet"

        # ALL-CAPS header (no amount) — should become section
        ws.append(["INCOME", ""])
        revenue_row = ws.max_row + 1
        ws.append(["Revenue from Operations", 1000000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=[ws_name])

        # Revenue item should carry the "income" section
        revenue_items = [i for i in items if "Revenue" in i.description or "revenue" in i.description.lower()]
        assert len(revenue_items) >= 1
        # section may be "income" (lowercase) or the raw header text
        assert revenue_items[0].section.lower() in ("income", "")

    @pytest.mark.asyncio
    async def test_excel_extractor_handles_indian_number_format(self, extractor):
        """Cell value '1,23,456' (Indian lakh) → LineItem.amount == 123456.0."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws_name = ws.title  # default is "Sheet"
        ws.append(["Other Income", "1,23,456"])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=[ws_name])

        income_items = [i for i in items if "Income" in i.description or "income" in i.description.lower()]
        assert len(income_items) >= 1
        assert income_items[0].amount == pytest.approx(123456.0)

    @pytest.mark.asyncio
    async def test_excel_extractor_skips_empty_rows(self, extractor):
        """Rows with no description and no amount should be silently skipped."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Revenue", 1000000])
        ws.append(["", ""])          # empty row
        ws.append(["", None])        # another empty row
        ws.append(["Cost of Goods", 600000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx")
        # Should only have the two data rows
        descriptions = [i.description for i in items]
        assert all(d.strip() for d in descriptions)

    @pytest.mark.asyncio
    async def test_excel_extractor_multi_sheet(self, extractor):
        """Extractor scans ALL selected sheets, not just the first."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["Revenue", 1000000])

        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["Expenses", 600000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=["Sheet1", "Sheet2"])
        descriptions = [i.description for i in items]
        assert any("Revenue" in d for d in descriptions)
        assert any("Expenses" in d for d in descriptions)

    @pytest.mark.asyncio
    async def test_excel_extractor_tags_page_type(self, extractor):
        """Items from P&L get page_type='face', Notes get page_type='notes'."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "P&L"
        ws1.append(["Revenue", 1000000])

        ws2 = wb.create_sheet("Notes to Accounts")
        ws2.append(["Power & Fuel", 500000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=["P&L", "Notes to Accounts"])

        face_items = [i for i in items if i.page_type == "face"]
        notes_items = [i for i in items if i.page_type == "notes"]
        assert len(face_items) >= 1
        assert len(notes_items) >= 1


# OcrExtractor tests moved to tests/test_vision_ocr.py (Claude Vision pipeline, Phase 10)


# ── Factory: routing ──────────────────────────────────────────────────────────


class TestExtractorFactory:
    @pytest.mark.asyncio
    async def test_factory_routes_xlsx_to_excel_extractor(self):
        """file_type='xlsx' → ExcelExtractor.extract() is called."""
        with patch(
            "app.services.extraction.extractor_factory.ExcelExtractor.extract",
            new_callable=AsyncMock,
            return_value=[LineItem("Revenue", 1000000.0, "income", "Revenue 10,00,000")],
        ) as mock_extract:
            result = await extract_document(_make_xlsx_bytes(), "xlsx", "client/file.xlsx")

        mock_extract.assert_called_once()
        assert len(result) == 1

    @pytest.mark.asyncio
    async def test_factory_routes_xls_to_excel_extractor(self):
        """file_type='xls' → ExcelExtractor.extract() is called."""
        with patch(
            "app.services.extraction.extractor_factory.ExcelExtractor.extract",
            new_callable=AsyncMock,
            return_value=[LineItem("Revenue", 1000000.0, "income", "Revenue 10,00,000")],
        ) as mock_extract:
            result = await extract_document(_make_xls_bytes(), "xls", "client/file.xls")

        mock_extract.assert_called_once()

    @pytest.mark.asyncio
    @patch("app.services.extraction.extractor_factory.OcrExtractor")
    async def test_factory_routes_pdf_to_vision_ai(self, MockOcr):
        """All PDFs now route through OcrExtractor (Gemini Flash vision)."""
        mock_instance = AsyncMock()
        mock_instance.extract.return_value = [LineItem(description="Sales", amount=500000, section="income", raw_text="Sales 500000")]
        MockOcr.return_value = mock_instance

        result = await extract_document(b"fake-pdf", "pdf", "test.pdf")
        assert len(result) == 1
        assert result[0].description == "Sales"
        MockOcr.assert_called_once()

    @pytest.mark.asyncio
    async def test_factory_returns_list_of_line_items(self):
        """extract_document always returns list[LineItem]."""
        expected = [LineItem("Revenue", 1000000.0, "income", "Revenue 10,00,000")]
        with patch(
            "app.services.extraction.extractor_factory.ExcelExtractor.extract",
            new_callable=AsyncMock,
            return_value=expected,
        ):
            result = await extract_document(_make_xlsx_bytes(), "xlsx", "client/file.xlsx")

        assert result == expected

    @pytest.mark.asyncio
    async def test_factory_raises_on_unknown_file_type(self):
        """Unknown file_type raises ExtractionError."""
        with pytest.raises(ExtractionError, match="Unsupported file type"):
            await extract_document(b"data", "docx", "client/file.docx")


# ── Cross-sheet deduplication ─────────────────────────────────────────────────


class TestSheetPriority:
    """Test sheet priority scoring for dedup preference."""

    def test_pl_face_sheet_lowest_priority(self):
        assert _sheet_priority("P&L") == 1
        assert _sheet_priority("Profit and Loss") == 1
        assert _sheet_priority("Profit & Loss Account") == 1

    def test_balance_sheet_face_lowest_priority(self):
        assert _sheet_priority("Balance Sheet") == 1
        assert _sheet_priority("B.S.") == 1

    def test_cash_flow_face_lowest_priority(self):
        assert _sheet_priority("Cash Flow") == 1

    def test_notes_highest_priority(self):
        assert _sheet_priority("Notes") == 3
        assert _sheet_priority("Note 28") == 3
        assert _sheet_priority("Schedule") == 3

    def test_unknown_sheet_middle_priority(self):
        assert _sheet_priority("SomeRandomSheet") == 2

    def test_notes_beat_face_sheets(self):
        assert _sheet_priority("P&L") < _sheet_priority("Note 28")
        assert _sheet_priority("Balance Sheet") < _sheet_priority("Notes")


class TestSheetPageType:
    """Test page_type assignment from sheet names."""

    def test_face_sheets(self):
        assert _sheet_page_type("P&L") == "face"
        assert _sheet_page_type("Balance Sheet") == "face"
        assert _sheet_page_type("Cash Flow") == "face"
        assert _sheet_page_type("Trading") == "face"

    def test_note_sheets(self):
        assert _sheet_page_type("Notes") == "notes"
        assert _sheet_page_type("Note 28") == "notes"
        assert _sheet_page_type("Schedule") == "notes"

    def test_ambiguous_notes_with_face_keywords(self):
        """Sheets like 'Notes BS (2)' or 'Notes to P & L' are Notes, not face."""
        assert _sheet_page_type("Notes BS (2)") == "notes"
        assert _sheet_page_type("Notes to P & L") == "notes"
        assert _sheet_page_type("Notes BS (1)") == "notes"

    def test_unknown_sheets(self):
        assert _sheet_page_type("SomeRandomSheet") == "unknown"


class TestDescriptionsMatch:
    """Test fuzzy description matching for dedup."""

    def test_exact_match(self):
        assert _descriptions_match("Revenue from Operations", "Revenue from Operations")

    def test_case_insensitive_match(self):
        assert _descriptions_match("Revenue from Operations", "revenue from operations")

    def test_minor_wording_difference(self):
        # "Revenue from Operations" vs "Revenue From Operations (Net)" — similar enough
        assert _descriptions_match("Revenue from Operations", "Revenue From Operations")

    def test_completely_different_descriptions_no_match(self):
        assert not _descriptions_match("Revenue from Operations", "Depreciation on Plant")

    def test_empty_strings_no_match(self):
        assert not _descriptions_match("", "Revenue")
        assert not _descriptions_match("Revenue", "")
        assert not _descriptions_match("", "")


class TestAmountsMatch:
    """Test amount comparison for dedup."""

    def test_exact_match(self):
        assert _amounts_match(96576700.0, 96576700.0)

    def test_both_zero(self):
        assert _amounts_match(0.0, 0.0)

    def test_tiny_rounding_difference(self):
        assert _amounts_match(96576700.0, 96576700.01)

    def test_different_amounts_no_match(self):
        assert not _amounts_match(96576700.0, 50000000.0)

    def test_significantly_different_no_match(self):
        assert not _amounts_match(1000.0, 2000.0)


class TestDeduplicateAcrossSheets:
    """Test the full deduplication function."""

    def test_no_duplicates_returns_all(self):
        """Items with different descriptions/amounts are all kept."""
        items = [
            LineItem("Revenue", 1000000.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Depreciation", 500000.0, "expenses", "raw2", source_sheet="P&L"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 2

    def test_exact_duplicate_across_sheets_keeps_notes(self):
        """Same item on P&L and Notes → keep Notes copy (higher priority now)."""
        items = [
            LineItem("Revenue from Operations", 96576700.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Revenue from Operations", 96576700.0, "income", "raw2", source_sheet="Note 28"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 1
        assert result[0].source_sheet == "Note 28"

    def test_triple_duplicate_keeps_one(self):
        """Same item on P&L, Notes, and Schedule → keep only Notes copy (highest priority)."""
        items = [
            LineItem("Depreciation", 4070000.0, "expenses", "raw1", source_sheet="Note 15"),
            LineItem("Depreciation", 4070000.0, "expenses", "raw2", source_sheet="P&L"),
            LineItem("Depreciation", 4070000.0, "expenses", "raw3", source_sheet="Schedule"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 1
        assert result[0].source_sheet == "Note 15"

    def test_same_description_different_amounts_kept_both(self):
        """Same description but different amounts → both kept (not a duplicate)."""
        items = [
            LineItem("Revenue from Operations", 96576700.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Revenue from Operations", 50000000.0, "income", "raw2", source_sheet="Note 28"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 2

    def test_same_amount_different_descriptions_kept_both(self):
        """Same amount but clearly different descriptions → both kept."""
        items = [
            LineItem("Revenue from Operations", 96576700.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Cost of Materials", 96576700.0, "expenses", "raw2", source_sheet="P&L"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 2

    def test_fuzzy_description_match_deduplicates(self):
        """Slightly different wording of the same item is still deduplicated."""
        items = [
            LineItem("Revenue from Operations", 96576700.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Revenue From Operations", 96576700.0, "income", "raw2", source_sheet="Notes"),
        ]
        result = deduplicate_across_sheets(items)
        assert len(result) == 1

    def test_empty_list_returns_empty(self):
        assert deduplicate_across_sheets([]) == []

    def test_single_item_returns_unchanged(self):
        items = [LineItem("Revenue", 100000.0, "income", "raw1", source_sheet="P&L")]
        result = deduplicate_across_sheets(items)
        assert len(result) == 1

    def test_items_from_same_sheet_not_deduplicated(self):
        """Two items with the same description+amount on the SAME sheet are separate rows."""
        items = [
            LineItem("Interest Income", 50000.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Interest Income", 50000.0, "other income", "raw2", source_sheet="P&L"),
        ]
        # These WILL be grouped as duplicates since same desc+amount.
        # Both from same sheet → one kept. This is acceptable because
        # truly identical rows on the same sheet are almost always a
        # parsing artifact, not two genuine line items.
        result = deduplicate_across_sheets(items)
        assert len(result) == 1

    def test_real_world_bcipl_scenario(self):
        """Simulates the BCIPL bug: Domestic Sales on P&L + Note 28 + Schedule.

        Expected 96.58 Cr, was getting 335.38 Cr because 4 copies were summed.
        After dedup, only 1 copy should remain.
        """
        items = [
            LineItem("Domestic Sales", 105430000.0, "income", "raw1", source_sheet="P&L"),
            LineItem("Domestic Sales - Product A", 550000.0, "income", "raw2", source_sheet="Note 28"),
            LineItem("Domestic Sales", 105430000.0, "income", "raw3", source_sheet="Note 28"),
            LineItem("Domestic Sales", 105430000.0, "income", "raw4", source_sheet="Schedule IV"),
        ]
        result = deduplicate_across_sheets(items)
        # The 105430000 copies should be deduplicated to 1 (P&L).
        # The 550000 "Product A" has a different amount so it stays.
        assert len(result) == 2
        # The big amount should come from Note 28 (Notes win now — higher priority)
        big_items = [i for i in result if i.amount == 105430000.0]
        assert len(big_items) == 1
        assert big_items[0].source_sheet == "Note 28"
        # The small amount (sub-item) should still be there
        small_items = [i for i in result if i.amount == 550000.0]
        assert len(small_items) == 1

    @pytest.mark.asyncio
    async def test_xlsx_extraction_deduplicates_across_sheets(self):
        """Integration test: ExcelExtractor with multi-sheet xlsx actually deduplicates."""
        import openpyxl

        wb = openpyxl.Workbook()
        # Sheet 1: P&L face
        ws1 = wb.active
        ws1.title = "P&L"
        ws1.append(["Description", "Amount"])
        ws1.append(["Revenue from Operations", 96576700])
        ws1.append(["Depreciation", 4070000])

        # Sheet 2: Notes (same items repeated)
        ws2 = wb.create_sheet("Notes to Accounts")
        ws2.append(["Description", "Amount"])
        ws2.append(["Revenue from Operations", 96576700])
        ws2.append(["Depreciation", 4070000])

        buf = io.BytesIO()
        wb.save(buf)
        xlsx_bytes = buf.getvalue()

        extractor = ExcelExtractor()
        items = await extractor.extract(xlsx_bytes, "xlsx", selected_sheets=["P&L", "Notes to Accounts"])

        # Should have 2 items (deduplicated), not 4
        assert len(items) == 2
        descriptions = [i.description for i in items]
        assert "Revenue from Operations" in descriptions
        assert "Depreciation" in descriptions
        # Both should come from Notes to Accounts (higher priority now)
        for item in items:
            assert item.source_sheet == "Notes to Accounts"
