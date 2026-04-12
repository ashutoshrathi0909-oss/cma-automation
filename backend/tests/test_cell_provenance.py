# backend/tests/test_cell_provenance.py
"""Phase 1: Per-Cell Audit Trail — TDD RED phase.

ALL tests in this file MUST FAIL before Phase 1 implementation begins.
After implementation, ALL tests MUST PASS.

Tests cover:
  1. _fill_data_cells returns provenance records (pure layer)
  2. _fetch_classified_data includes line_item_id + classification_id
  3. generate() inserts provenance to cell_provenance table
  4. GET /api/cma-reports/{id}/provenance endpoint
  5. Provenance data integrity (amounts, row/col mapping)
"""

from __future__ import annotations

import openpyxl
import pytest
from unittest.mock import MagicMock, patch, call
from openpyxl.utils import column_index_from_string

from app.services.excel_generator import ExcelGenerator


# ── Helpers ─────────────────────────────────────────────────────────────────

def _make_ws():
    """Return a minimal workbook + its INPUT SHEET worksheet."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "INPUT SHEET"
    return wb, ws


def _make_generator():
    """ExcelGenerator with mocked service (pure layer tests don't need DB)."""
    return ExcelGenerator(service=MagicMock(), template_path="/fake/template.xlsm")


# ══════════════════════════════════════════════════════════════════════════════
# 1. Pure layer: _fill_data_cells returns provenance records
# ══════════════════════════════════════════════════════════════════════════════


class TestFillDataCellsProvenance:
    """_fill_data_cells must return a list of provenance dicts alongside filling cells."""

    def test_returns_list_of_provenance_records(self):
        """_fill_data_cells returns a list (not None) of provenance dicts."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Raw Materials Consumed ( Indigenous)",
                "cma_row": 42,
                "financial_year": 2024,
                "amount": 500000,
                "document_id": "doc-001",
                "line_item_id": "li-001",
                "classification_id": "clf-001",
                "source_text": "Raw Materials Indigenous",
            }
        ]
        year_map = {2024: "B"}

        result = gen._fill_data_cells(ws, cell_data, year_map=year_map)

        # Currently _fill_data_cells returns None — this MUST fail
        assert isinstance(result, list), "_fill_data_cells must return a list of provenance records"
        assert len(result) == 1

    def test_provenance_contains_required_fields(self):
        """Each provenance record has all required fields."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Wages",
                "cma_row": 45,
                "financial_year": 2023,
                "amount": 300000,
                "document_id": "doc-002",
                "line_item_id": "li-002",
                "classification_id": "clf-002",
                "source_text": "Staff Wages",
            }
        ]
        year_map = {2023: "B"}

        result = gen._fill_data_cells(ws, cell_data, year_map=year_map)
        assert isinstance(result, list)
        prov = result[0]

        # All required fields must be present
        assert prov["cma_row"] == 45
        assert prov["cma_column"] == "B"
        assert prov["financial_year"] == 2023
        assert prov["line_item_id"] == "li-002"
        assert prov["classification_id"] == "clf-002"
        assert prov["source_text"] == "Staff Wages"
        assert prov["raw_amount"] == 300000
        assert prov["document_id"] == "doc-002"

    def test_provenance_converted_amount_uses_doc_divisor(self):
        """Converted amount reflects per-document unit conversion."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Wages",
                "cma_row": 45,
                "financial_year": 2024,
                "amount": 500000,
                "document_id": "doc-003",
                "line_item_id": "li-003",
                "classification_id": "clf-003",
                "source_text": "Wages",
            }
        ]
        year_map = {2024: "B"}
        doc_divisors = {"doc-003": 100000}  # rupees -> lakhs

        result = gen._fill_data_cells(
            ws, cell_data, year_map=year_map, doc_divisors=doc_divisors
        )
        assert isinstance(result, list)
        prov = result[0]

        assert prov["raw_amount"] == 500000
        assert prov["converted_amount"] == 5.0  # 500000 / 100000

    def test_skipped_items_not_in_provenance(self):
        """Items with no valid row or no valid column are not in provenance."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Unknown Field",
                "cma_row": 0,  # invalid row
                "financial_year": 2024,
                "amount": 100000,
                "document_id": "doc-004",
                "line_item_id": "li-004",
                "classification_id": "clf-004",
                "source_text": "Unknown",
            }
        ]
        year_map = {2024: "B"}

        result = gen._fill_data_cells(ws, cell_data, year_map=year_map)
        assert isinstance(result, list)
        assert len(result) == 0, "Skipped items must not appear in provenance"

    def test_multiple_items_same_cell_all_tracked(self):
        """When multiple items map to the same cell, each gets a provenance record."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Wages",
                "cma_row": 45,
                "financial_year": 2024,
                "amount": 300000,
                "document_id": "doc-005",
                "line_item_id": "li-005a",
                "classification_id": "clf-005a",
                "source_text": "Wages - Unit A",
            },
            {
                "cma_field_name": "Wages",
                "cma_row": 45,
                "financial_year": 2024,
                "amount": 200000,
                "document_id": "doc-005",
                "line_item_id": "li-005b",
                "classification_id": "clf-005b",
                "source_text": "Wages - Unit B",
            },
        ]
        year_map = {2024: "B"}

        result = gen._fill_data_cells(ws, cell_data, year_map=year_map)
        assert isinstance(result, list)
        assert len(result) == 2, "Both items must have provenance records"
        assert {r["line_item_id"] for r in result} == {"li-005a", "li-005b"}


# ══════════════════════════════════════════════════════════════════════════════
# 2. _fetch_classified_data must include line_item_id + classification_id
# ══════════════════════════════════════════════════════════════════════════════


class TestFetchClassifiedDataProvenance:
    """_fetch_classified_data must return line_item_id and classification_id per item."""

    def test_cell_data_includes_line_item_id(self):
        """Each cell_data dict must have a 'line_item_id' key."""
        service = MagicMock()
        gen = ExcelGenerator(service=service, template_path="/fake/template.xlsm")

        # Mock extracted_line_items query
        service.table("extracted_line_items").select.return_value \
            .eq.return_value.range.return_value.execute.return_value.data = [
                {"id": "li-100", "document_id": "doc-100", "amount": 500000}
            ]
        # Second page returns empty (pagination stop)
        service.table("extracted_line_items").select.return_value \
            .eq.return_value.range.return_value.execute.return_value.data = [
                {"id": "li-100", "document_id": "doc-100", "amount": 500000}
            ]

        # Mock classifications query
        service.table("classifications").select.return_value \
            .in_.return_value.eq.return_value.execute.return_value.data = [
                {
                    "id": "clf-100",
                    "line_item_id": "li-100",
                    "cma_field_name": "Wages",
                    "cma_row": 45,
                }
            ]

        result = gen._fetch_classified_data(
            document_ids=["doc-100"],
            doc_years={"doc-100": 2024},
        )

        assert len(result) >= 1
        item = result[0]
        assert "line_item_id" in item, "cell_data must include line_item_id for provenance"
        assert item["line_item_id"] == "li-100"

    def test_cell_data_includes_classification_id(self):
        """Each cell_data dict must have a 'classification_id' key."""
        service = MagicMock()
        gen = ExcelGenerator(service=service, template_path="/fake/template.xlsm")

        service.table("extracted_line_items").select.return_value \
            .eq.return_value.range.return_value.execute.return_value.data = [
                {"id": "li-200", "document_id": "doc-200", "amount": 100000}
            ]
        service.table("classifications").select.return_value \
            .in_.return_value.eq.return_value.execute.return_value.data = [
                {
                    "id": "clf-200",
                    "line_item_id": "li-200",
                    "cma_field_name": "Rent , Rates and Taxes",
                    "cma_row": 68,
                }
            ]

        result = gen._fetch_classified_data(
            document_ids=["doc-200"],
            doc_years={"doc-200": 2024},
        )

        assert len(result) >= 1
        item = result[0]
        assert "classification_id" in item, "cell_data must include classification_id for provenance"
        assert item["classification_id"] == "clf-200"


# ══════════════════════════════════════════════════════════════════════════════
# 3. generate() must insert provenance into cell_provenance table
# ══════════════════════════════════════════════════════════════════════════════


class TestGenerateInsertProvenance:
    """ExcelGenerator.generate() must batch-insert provenance to cell_provenance table."""

    @patch("app.services.excel_generator.openpyxl.load_workbook")
    def test_generate_inserts_provenance_records(self, mock_load_wb):
        """After filling the workbook, provenance is batch-inserted to cell_provenance."""
        # Per-table mocks (service.table() returns same mock regardless of arg
        # with plain MagicMock, so we route via side_effect)
        tbl_reports = MagicMock()
        tbl_clients = MagicMock()
        tbl_documents = MagicMock()
        tbl_line_items = MagicMock()
        tbl_classifications = MagicMock()
        tbl_provenance = MagicMock()

        table_map = {
            "cma_reports": tbl_reports,
            "clients": tbl_clients,
            "documents": tbl_documents,
            "extracted_line_items": tbl_line_items,
            "classifications": tbl_classifications,
            "cell_provenance": tbl_provenance,
        }
        service = MagicMock()
        service.table.side_effect = lambda name: table_map.get(name, MagicMock())
        gen = ExcelGenerator(service=service, template_path="/fake/template.xlsm")

        # Mock report fetch
        tbl_reports.select.return_value \
            .eq.return_value.single.return_value.execute.return_value.data = {
                "id": "rpt-001",
                "client_id": "client-001",
                "document_ids": ["doc-001"],
                "cma_output_unit": "lakhs",
                "status": "generating",
            }
        # Mock client name
        tbl_clients.select.return_value \
            .eq.return_value.single.return_value.execute.return_value.data = {
                "name": "Test Co"
            }
        # Mock documents
        tbl_documents.select.return_value \
            .in_.return_value.execute.return_value.data = [
                {"id": "doc-001", "financial_year": 2024, "nature": "audited", "source_unit": "rupees"}
            ]
        # Mock line items
        tbl_line_items.select.return_value \
            .eq.return_value.range.return_value.execute.return_value.data = [
                {"id": "li-001", "document_id": "doc-001", "amount": 500000}
            ]
        # Mock classifications
        tbl_classifications.select.return_value \
            .in_.return_value.eq.return_value.execute.return_value.data = [
                {
                    "id": "clf-001",
                    "line_item_id": "li-001",
                    "cma_field_name": "Wages",
                    "cma_row": 45,
                }
            ]

        # Mock openpyxl workbook
        mock_wb = MagicMock()
        mock_ws = MagicMock()
        mock_ws.cell.return_value = MagicMock(value=None)
        mock_wb.__getitem__ = MagicMock(return_value=mock_ws)
        mock_load_wb.return_value = mock_wb

        # Mock save/upload/cleanup
        gen._save_upload_cleanup = MagicMock(return_value="cma_reports/rpt-001/output.xlsm")

        # Run generate
        gen.generate(report_id="rpt-001", user_id="user-001")

        # Verify cell_provenance insert was called
        assert tbl_provenance.insert.called, (
            "generate() must insert provenance records into cell_provenance table"
        )


# ══════════════════════════════════════════════════════════════════════════════
# 4. GET /api/cma-reports/{id}/provenance endpoint
# ══════════════════════════════════════════════════════════════════════════════


class TestProvenanceEndpoint:
    """Provenance API endpoint must exist and return correct data."""

    def test_endpoint_exists_and_returns_200(self, admin_client):
        """GET /api/cma-reports/{id}/provenance returns 200."""
        from unittest.mock import patch

        mock_service = MagicMock()
        # Mock _get_owned_report's query
        mock_service.table("cma_reports").select.return_value \
            .eq.return_value.single.return_value.execute.return_value.data = {
                "id": "rpt-001", "created_by": "admin-uuid-0001",
            }
        # Mock provenance query
        mock_service.table("cell_provenance").select.return_value \
            .eq.return_value.eq.return_value.eq.return_value \
            .execute.return_value.data = []

        with patch("app.routers.cma_reports.get_service_client", return_value=mock_service):
            resp = admin_client.get(
                "/api/cma-reports/rpt-001/provenance",
                params={"row": 42, "column": "B"},
            )
        assert resp.status_code == 200, (
            f"Expected 200 but got {resp.status_code}. "
            "The /provenance endpoint must exist on cma_reports router."
        )

    def test_returns_provenance_records_for_cell(self, admin_client):
        """Endpoint returns provenance records filtered by row + column."""
        from unittest.mock import patch

        provenance_record = {
            "id": "prov-001",
            "cma_report_id": "rpt-001",
            "cma_row": 45,
            "cma_column": "B",
            "financial_year": 2024,
            "source_text": "Wages & Salaries",
            "raw_amount": 500000.0,
            "converted_amount": 5.0,
            "document_id": "doc-001",
        }
        mock_service = MagicMock()
        mock_service.table("cma_reports").select.return_value \
            .eq.return_value.single.return_value.execute.return_value.data = {
                "id": "rpt-001", "created_by": "admin-uuid-0001",
            }
        mock_service.table("cell_provenance").select.return_value \
            .eq.return_value.eq.return_value.eq.return_value \
            .execute.return_value.data = [provenance_record]

        with patch("app.routers.cma_reports.get_service_client", return_value=mock_service):
            resp = admin_client.get(
                "/api/cma-reports/rpt-001/provenance",
                params={"row": 45, "column": "B"},
            )
        assert resp.status_code == 200
        data = resp.json()
        assert isinstance(data, list), "Response must be a list of provenance records"
        assert len(data) == 1
        record = data[0]
        required_fields = {"cma_row", "cma_column", "source_text"}
        assert required_fields.issubset(record.keys()), (
            f"Provenance record missing fields: {required_fields - record.keys()}"
        )

    def test_requires_row_and_column_params(self, admin_client):
        """Endpoint returns 422 when row or column params are missing."""
        resp = admin_client.get("/api/cma-reports/rpt-001/provenance")
        assert resp.status_code == 422, "Must require row and column query params"


# ══════════════════════════════════════════════════════════════════════════════
# 5. Provenance data integrity
# ══════════════════════════════════════════════════════════════════════════════


class TestProvenanceDataIntegrity:
    """Provenance records must accurately reflect the data that was written to Excel."""

    def test_provenance_row_matches_cell_written(self):
        """Provenance cma_row matches the row where the cell value was written."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {
                "cma_field_name": "Rent , Rates and Taxes",
                "cma_row": 68,
                "financial_year": 2024,
                "amount": 120000,
                "document_id": "doc-int-1",
                "line_item_id": "li-int-1",
                "classification_id": "clf-int-1",
                "source_text": "Rent Expense",
            }
        ]
        year_map = {2024: "C"}

        provenance = gen._fill_data_cells(ws, cell_data, year_map=year_map)
        assert isinstance(provenance, list)

        # The cell at (68, C=3) should have been written
        written_value = ws.cell(row=68, column=3).value
        assert written_value == 120000

        # And provenance should point to the same cell
        assert provenance[0]["cma_row"] == 68
        assert provenance[0]["cma_column"] == "C"

    def test_provenance_count_matches_items_processed(self):
        """Number of provenance records equals number of items successfully processed."""
        _, ws = _make_ws()
        gen = _make_generator()

        cell_data = [
            {"cma_field_name": "Wages", "cma_row": 45, "financial_year": 2024,
             "amount": 100000, "document_id": "d1", "line_item_id": "l1",
             "classification_id": "c1", "source_text": "W1"},
            {"cma_field_name": "Wages", "cma_row": 45, "financial_year": 2024,
             "amount": 200000, "document_id": "d1", "line_item_id": "l2",
             "classification_id": "c2", "source_text": "W2"},
            {"cma_field_name": "Bad", "cma_row": 0, "financial_year": 2024,
             "amount": 50000, "document_id": "d1", "line_item_id": "l3",
             "classification_id": "c3", "source_text": "Skip me"},
        ]
        year_map = {2024: "B"}

        provenance = gen._fill_data_cells(ws, cell_data, year_map=year_map)
        assert isinstance(provenance, list)
        assert len(provenance) == 2, "2 valid items processed, 1 skipped"
