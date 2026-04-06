"""DeepSeek AI Golden Test Suite — adversarial stress test.

Every item has a known correct answer from CA-verified decisions.
Many are curve balls where the AI's naive suggestion was WRONG.

Run mock tests:     pytest tests/test_deepseek_golden.py -v -k "not Live"
Run live API test:  pytest tests/test_deepseek_golden.py -v -k "Live"
"""

from __future__ import annotations

import os
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pytest
from openpyxl.utils import column_index_from_string

from tests.fixtures.golden_dataset import (
    EXPECTED_FORMULAS,
    EXPECTED_PLAIN_VALUES,
    GOLDEN_CLIENT_ID,
    GOLDEN_ITEMS,
    GOLDEN_YEAR,
    get_ai_only_items,
    get_industry_pairs,
    make_cell_data_item,
    make_pipeline_item,
)


# ── Helpers ────────────────────────────────────────────────────────────────


def _make_ai_result(item):
    from app.services.classification.ai_classifier import AIClassificationResult
    return AIClassificationResult(
        cma_field_name=item["expected_cma_field"],
        cma_row=item["expected_cma_row"],
        cma_sheet="input_sheet",
        broad_classification=None,
        confidence=0.92,
        is_doubt=False,
        doubt_reason=None,
        alternatives=[],
        classification_method="scoped_v3",
    )


def _make_doubt():
    from app.services.classification.ai_classifier import AIClassificationResult
    return AIClassificationResult(
        cma_field_name="UNCLASSIFIED", cma_row=0, cma_sheet="input_sheet",
        broad_classification=None, confidence=0.3, is_doubt=True,
        doubt_reason="No match in golden test", alternatives=[],
        classification_method="scoped_v3",
    )


def _run_pipeline(items=None):
    from app.services.classification.pipeline import ClassificationPipeline
    if items is None:
        items = GOLDEN_ITEMS
    os.environ.setdefault("PROJECT_ROOT", str(Path(__file__).parents[1].parent))

    # Key by (description, industry_type) to handle items like T04/T19 that
    # have the same text but different expected rows per industry.
    ai_answers = {
        (i["description"], i["industry_type"]): _make_ai_result(i)
        for i in get_ai_only_items()
    }

    class MockAI:
        def classify(self, **kw):
            key = (kw.get("raw_text"), kw.get("industry_type"))
            return ai_answers.get(key, _make_doubt())
        def classify_sync(self, **kw):
            key = (kw.get("raw_text"), kw.get("industry_type"))
            return ai_answers.get(key, _make_doubt())
        def set_learned_cache(self, mappings):
            pass

    with patch("app.services.classification.pipeline.ScopedClassifier", MockAI), \
         patch("app.services.classification.pipeline.get_settings") as ms:
        ms.return_value = MagicMock(classifier_mode="scoped", openrouter_api_key="test")
        pipeline = ClassificationPipeline()
        results = []
        for item in items:
            r = pipeline.classify_item(
                item=make_pipeline_item(item),
                client_id=GOLDEN_CLIENT_ID,
                industry_type=item["industry_type"],
                document_type=item["document_type"],
                financial_year=item["financial_year"],
            )
            r["_id"] = item["id"]
            r["_expected_row"] = item["expected_cma_row"]
            r["_expected_field"] = item["expected_cma_field"]
            r["_twist"] = item.get("twist", "")
            results.append(r)
        return results


# Known pipeline gaps — items where the AI-only pipeline still misclassifies.
# As of April 2026 the regex/fuzzy tiers were removed, so old regex misroute
# gaps (T12, T24) are resolved — AI handles them correctly.
KNOWN_PIPELINE_GAPS: set[str] = set()


# ═══ Test Class 1: All items classified correctly ═══

class TestGoldenClassification:

    def test_all_items_correct_row(self):
        results = _run_pipeline()
        failures = []
        for r in results:
            if r["_id"] in KNOWN_PIPELINE_GAPS:
                continue
            if r["cma_row"] != r["_expected_row"]:
                failures.append(
                    f"  {r['_id']}: expected R{r['_expected_row']}, "
                    f"got R{r['cma_row']} ({r['cma_field_name']})\n"
                    f"    Twist: {r['_twist']}"
                )
        assert not failures, f"\n{len(failures)} errors:\n" + "\n".join(failures)

    def test_known_gaps_documented(self):
        """Verify known pipeline gaps still fail (remove from set when fixed)."""
        results = _run_pipeline()
        for r in results:
            if r["_id"] in KNOWN_PIPELINE_GAPS:
                assert r["cma_row"] != r["_expected_row"], (
                    f"{r['_id']} is now CORRECT — remove from KNOWN_PIPELINE_GAPS!"
                )

    def test_no_doubts(self):
        results = _run_pipeline()
        doubts = [r for r in results if r.get("is_doubt")]
        assert not doubts, "\n".join(f"  {r['_id']}: {r.get('doubt_reason')}" for r in doubts)

    def test_count_matches(self):
        assert len(_run_pipeline()) == len(GOLDEN_ITEMS)


# ═══ Test Class 2: Industry routing ═══

class TestGoldenIndustryRouting:

    def test_staff_welfare_mfg_vs_trading(self):
        for mfg, trd in get_industry_pairs():
            mr = _run_pipeline([mfg])[0]
            tr = _run_pipeline([trd])[0]
            assert mr["cma_row"] == mfg["expected_cma_row"]
            assert tr["cma_row"] == trd["expected_cma_row"]
            assert mr["cma_row"] != tr["cma_row"], "Same item must route differently by industry"

    def test_bonus_mfg_r45_trading_r67(self):
        # T05 has "bonus" in manufacturing context → R45 via regex
        # T20 has "bonus" in trading context → R67
        t20 = next(i for i in GOLDEN_ITEMS if i["id"] == "T20")
        r = _run_pipeline([t20])[0]
        assert r["cma_row"] == 67, f"Trading bonus: expected R67, got R{r['cma_row']}"


# ═══ Test Class 3: CA Override edge cases ═══

class TestGoldenCAOverrides:

    def test_gratuity_r45(self):
        r = _run_pipeline([GOLDEN_ITEMS[0]])[0]
        assert r["cma_row"] == 45

    def test_liquidated_damages_r71(self):
        item = next(i for i in GOLDEN_ITEMS if i["id"] == "T08")
        assert _run_pipeline([item])[0]["cma_row"] == 71

    def test_loan_processing_fee_r85(self):
        item = next(i for i in GOLDEN_ITEMS if i["id"] == "T17")
        assert _run_pipeline([item])[0]["cma_row"] == 85

    def test_misc_expenses_r71_not_r75(self):
        item = next(i for i in GOLDEN_ITEMS if i["id"] == "T23")
        assert _run_pipeline([item])[0]["cma_row"] == 71

    def test_directors_remuneration_r73(self):
        item = next(i for i in GOLDEN_ITEMS if i["id"] == "T18")
        assert _run_pipeline([item])[0]["cma_row"] == 73


# ═══ Test Class 4: Formula output ═══

class TestGoldenFormulas:

    def _build_ws(self):
        from app.services.excel_generator import ExcelGenerator
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "INPUT SHEET"
        gen = ExcelGenerator(service=MagicMock(), template_path="/fake/template.xlsm")
        gen._fill_data_cells(ws, [make_cell_data_item(i) for i in GOLDEN_ITEMS], year_map={GOLDEN_YEAR: "B"})
        return ws

    def test_multi_item_formulas(self):
        ws = self._build_ws()
        col = column_index_from_string("B")
        for (row, _), expected in EXPECTED_FORMULAS.items():
            actual = ws.cell(row=row, column=col).value
            assert actual == expected, f"R{row}: expected '{expected}', got '{actual}'"

    def test_single_item_plain_values(self):
        ws = self._build_ws()
        col = column_index_from_string("B")
        for (row, _), expected in EXPECTED_PLAIN_VALUES.items():
            actual = ws.cell(row=row, column=col).value
            assert actual == expected, f"R{row}: expected {expected}, got '{actual}'"

    def test_no_sum_function(self):
        ws = self._build_ws()
        col = column_index_from_string("B")
        for (row, _), formula in EXPECTED_FORMULAS.items():
            val = str(ws.cell(row=row, column=col).value)
            assert "SUM" not in val.upper(), f"R{row}: must not use SUM()"

    def test_no_plus_minus_concatenation(self):
        """Negative amounts must show as subtraction, never +-."""
        ws = self._build_ws()
        col = column_index_from_string("B")
        for (row, _), formula in EXPECTED_FORMULAS.items():
            val = str(ws.cell(row=row, column=col).value)
            assert "+-" not in val, f"R{row}: got '+-' in '{val}'"

    def test_subtraction_present_where_expected(self):
        """Formulas with negative items must contain '-' (not just '+')."""
        ws = self._build_ws()
        col = column_index_from_string("B")
        # R45 has one negative (-15000), R22 has two negatives, R84 has one negative
        for row in [45, 22, 84]:
            val = str(ws.cell(row=row, column=col).value)
            assert "-" in val[1:], f"R{row}: expected subtraction in formula, got '{val}'"


# ═══ Test Class 5: Live DeepSeek API (Opt-in) ═══

