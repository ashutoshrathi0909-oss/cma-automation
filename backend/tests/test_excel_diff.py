# backend/tests/test_excel_diff.py
"""Phase 2: Excel diff service — TDD RED phase.

ALL tests MUST FAIL before implementation (module does not exist).
Zero API calls — pure Python file comparison.
"""

import openpyxl
import pytest


class TestDiffCmaFiles:
    """diff_cma_files compares AI-generated vs father-corrected CMA."""

    def test_import_succeeds(self):
        """Module must be importable."""
        from app.services.excel_diff import diff_cma_files  # noqa: F401

    def test_detects_changed_numeric_cell(self, tmp_path):
        """Diff detects when father changed a numeric cell value."""
        from app.services.excel_diff import diff_cma_files

        ai_wb = openpyxl.Workbook()
        ai_ws = ai_wb.active
        ai_ws.title = "INPUT SHEET"
        ai_ws.cell(row=42, column=2, value=500000)
        ai_path = tmp_path / "ai.xlsx"
        ai_wb.save(ai_path)

        father_wb = openpyxl.Workbook()
        father_ws = father_wb.active
        father_ws.title = "INPUT SHEET"
        father_ws.cell(row=42, column=2, value=400000)
        father_path = tmp_path / "father.xlsx"
        father_wb.save(father_path)

        diffs = diff_cma_files(str(ai_path), str(father_path))
        assert len(diffs) == 1
        assert diffs[0].cma_row == 42
        assert diffs[0].ai_value == 500000
        assert diffs[0].father_value == 400000

    def test_ignores_identical_cells(self, tmp_path):
        """Identical cells produce no diff."""
        from app.services.excel_diff import diff_cma_files

        for name in ("ai.xlsx", "father.xlsx"):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "INPUT SHEET"
            ws.cell(row=45, column=2, value=300000)
            wb.save(tmp_path / name)

        diffs = diff_cma_files(str(tmp_path / "ai.xlsx"), str(tmp_path / "father.xlsx"))
        assert len(diffs) == 0

    def test_ignores_formula_cells(self, tmp_path):
        """Cells with formulas (starting with =) are skipped."""
        from app.services.excel_diff import diff_cma_files

        ai_wb = openpyxl.Workbook()
        ai_ws = ai_wb.active
        ai_ws.title = "INPUT SHEET"
        ai_ws.cell(row=63, column=2, value="=SUM(B56)")
        ai_wb.save(tmp_path / "ai.xlsx")

        father_wb = openpyxl.Workbook()
        father_ws = father_wb.active
        father_ws.title = "INPUT SHEET"
        father_ws.cell(row=63, column=2, value="=SUM(B56,B49)")
        father_wb.save(tmp_path / "father.xlsx")

        diffs = diff_cma_files(str(tmp_path / "ai.xlsx"), str(tmp_path / "father.xlsx"))
        assert len(diffs) == 0, "Formula cells must be ignored"

    def test_detects_value_added_by_father(self, tmp_path):
        """Diff detects when father added a value where AI had nothing."""
        from app.services.excel_diff import diff_cma_files

        ai_wb = openpyxl.Workbook()
        ai_ws = ai_wb.active
        ai_ws.title = "INPUT SHEET"
        ai_wb.save(tmp_path / "ai.xlsx")

        father_wb = openpyxl.Workbook()
        father_ws = father_wb.active
        father_ws.title = "INPUT SHEET"
        father_ws.cell(row=67, column=3, value=250000)
        father_wb.save(tmp_path / "father.xlsx")

        diffs = diff_cma_files(str(tmp_path / "ai.xlsx"), str(tmp_path / "father.xlsx"))
        assert len(diffs) == 1
        assert diffs[0].ai_value is None
        assert diffs[0].father_value == 250000

    def test_numeric_tolerance(self, tmp_path):
        """Values within 0.01 tolerance are treated as equal."""
        from app.services.excel_diff import diff_cma_files

        ai_wb = openpyxl.Workbook()
        ai_ws = ai_wb.active
        ai_ws.title = "INPUT SHEET"
        ai_ws.cell(row=42, column=2, value=500000.005)
        ai_wb.save(tmp_path / "ai.xlsx")

        father_wb = openpyxl.Workbook()
        father_ws = father_wb.active
        father_ws.title = "INPUT SHEET"
        father_ws.cell(row=42, column=2, value=500000.01)
        father_wb.save(tmp_path / "father.xlsx")

        diffs = diff_cma_files(str(tmp_path / "ai.xlsx"), str(tmp_path / "father.xlsx"))
        assert len(diffs) == 0, "Within tolerance — should not produce a diff"


class TestCellDiffDataclass:
    """CellDiff must have the required fields."""

    def test_celldiff_fields(self):
        from app.services.excel_diff import CellDiff
        d = CellDiff(cma_row=42, cma_column="B", ai_value=500000, father_value=400000)
        assert d.cma_row == 42
        assert d.cma_column == "B"
        assert d.ai_value == 500000
        assert d.father_value == 400000
