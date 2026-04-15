"""Scan an INPUT SHEET for numeric fills in rows that must stay empty.

Uses DOCS/cma_cell_types.json (Phase 0 output) as the source of truth for
row classification. Prints a report of bad fills; exits 0 if clean, 1 if
any bad fills found.

Usage:
  python scripts/scan_bad_fills.py <path/to/output.xlsm>
  python scripts/scan_bad_fills.py --all   # scan every *.xlsm in CMA_Extraction_v2/**
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
CELL_TYPES_PATH = REPO / "DOCS" / "cma_cell_types.json"
FORBIDDEN_TYPES = {"WHITE_HEADER", "NOTE_ROW", "BLANK"}


def scan_file(xlsm_path: Path, cell_types: dict) -> list[tuple]:
    """Return list of (row, type, label, col_letter, value) bad fills."""
    try:
        wb = openpyxl.load_workbook(str(xlsm_path), keep_vba=True, data_only=False)
    except Exception as e:
        print(f"[ERROR] cannot open {xlsm_path}: {e}")
        return []
    if "INPUT SHEET" not in wb.sheetnames:
        return []
    ws = wb["INPUT SHEET"]
    bad = []
    for r_str, info in cell_types["rows"].items():
        if info["type"] not in FORBIDDEN_TYPES:
            continue
        r = int(r_str)
        for col in range(2, 9):  # B..H
            v = ws.cell(row=r, column=col).value
            if v in (None, "", 0):
                continue
            if isinstance(v, str) and v.startswith("="):
                continue
            if isinstance(v, (int, float)):
                bad.append((r, info["type"], info.get("label"), chr(64 + col), v))
    return bad


def main() -> int:
    p = argparse.ArgumentParser()
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument("path", nargs="?", type=Path, help="Path to an .xlsm file")
    g.add_argument("--all", action="store_true", help="Scan all .xlsm under CMA_Extraction_v2/")
    args = p.parse_args()

    cell_types = json.loads(CELL_TYPES_PATH.read_text(encoding="utf-8"))

    paths: list[Path]
    if args.all:
        root = REPO / "CMA_Extraction_v2"
        paths = [p for p in root.rglob("*.xlsm") if not p.name.startswith("~")]
    else:
        paths = [args.path]

    total_bad = 0
    for p in paths:
        rel = p.relative_to(REPO) if p.is_absolute() and REPO in p.parents else p
        bad = scan_file(p, cell_types)
        total_bad += len(bad)
        if bad:
            print(f"\n=== {rel}: {len(bad)} bad fills ===")
            for r, t, l, c, v in bad:
                print(f"  R{r:3} [{t:12}] [{c}] {l}  = {v}")
        else:
            print(f"{rel}: OK (0 bad fills)")

    print(f"\nTotal bad fills across {len(paths)} file(s): {total_bad}")
    return 0 if total_bad == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
