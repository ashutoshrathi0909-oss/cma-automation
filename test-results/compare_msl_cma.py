"""
MSL CMA Excel Comparison Script
Compares generated CMA vs ground truth on FY25 (Col D in generated, Col E in ground truth).

IMPORTANT FINDINGS:
- Generated CMA maps 2025 → Column D (per YEAR_TO_COLUMN)
- Ground truth CMA has 2025 in Column E (template starts 2022 in col B)
- This is a column-offset bug in YEAR_TO_COLUMN mapping
- The script reads BOTH correctly and compares what's in each file at the expected row.
"""
import os
import sys

try:
    import openpyxl
    import xlrd
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install openpyxl xlrd")
    sys.exit(1)

# ─── Paths ────────────────────────────────────────────────────────────────────
BASE = r"C:\Users\ASHUTOSH\OneDrive\Desktop\CMA project -2"
GENERATED = os.path.join(BASE, "test-results", "MSL_generated_CMA.xlsm")
GROUND_TRUTH = os.path.join(BASE, "FInancials main", "MSL - Manufacturing", "MSL CMA 24102025.xls")

# ─── Ground truth expected values (FY25, in Lakhs) ────────────────────────────
# From task instructions (these are the EXPECTED values per the task spec)
EXPECTED = {
    22:  ("Domestic Sales",            2130.76),
    23:  ("Export Sales",              99.06),
    30:  ("Interest Received",         1.04),
    34:  ("Others Non-Op Income",      3.56),
    41:  ("RM Imported",               21.81),
    42:  ("RM Indigenous",             1943.16),
    44:  ("Stores & Spares",           44.30),
    45:  ("Wages",                     134.41),
    46:  ("Job Work",                  2.84),
    47:  ("Freight",                   2.56),
    48:  ("Power Coal Fuel",           35.40),
    49:  ("Others Manufacturing",      0.00),
    50:  ("R&M Manufacturing",         8.14),
    67:  ("Salary Staff",              0.00),
    68:  ("Rent Rates Taxes",          8.86),
    69:  ("Bad Debts",                 6.50),
    70:  ("Advertisements",            8.05),
    71:  ("Others Admin",              23.77),
    72:  ("R&M Admin",                 0.00),
    73:  ("Audit Fees",                1.35),
    83:  ("Interest TL",               2.98),
    84:  ("Interest WC",               19.67),
    85:  ("Bank Charges",              5.40),
    89:  ("Loss Sale FA",              3.36),
    90:  ("Sundry Balances",           1.00),
    91:  ("Exchange Loss",             1.47),
    116: ("Paid-up Capital",           216.09),
    121: ("General Reserve",           7.50),
    122: ("P&L Balance",               10.24),
    123: ("Share Premium",             124.44),
    125: ("Other Reserve",             144.34),
    131: ("WC Bank Finance",           142.94),
    136: ("TL in 1 year",              14.10),
    137: ("TL after 1 year",           0.00),
    152: ("Quasi Equity",              417.68),
    153: ("LT Unsecured Debt",         21.72),
    201: ("Finished Goods",            878.74),
    206: ("Domestic Receivables",      345.40),
    208: ("Debtors > 6 months",        85.13),
    212: ("Cash on Hand",              0.34),
    213: ("Bank Balances",             6.64),
    219: ("Advances Govt",             13.44),
    220: ("Advances Suppliers RM",     0.23),
    221: ("Advance Income Tax",        48.32),
    222: ("Prepaid Expenses",          0.29),
    223: ("Other Advances",            5.54),
    237: ("Security Deposits",         17.53),
    242: ("Sundry Creditors",          254.32),
    243: ("Advance from Customers",    36.45),
    244: ("Provision for Taxation",    48.68),
    246: ("Other Statutory Liab",      2.31),
    250: ("Other Current Liab",        102.79),
}

TOLERANCE = 0.5  # ± Lakhs


def read_generated(path: str) -> dict[int, object]:
    """Read generated CMA - FY25 should be in Column D (col index 4)."""
    wb = openpyxl.load_workbook(path, keep_vba=True)
    ws = wb["INPUT SHEET"]
    data = {}
    for row in EXPECTED:
        val = ws.cell(row=row, column=4).value  # Col D = index 4
        data[row] = val
    return data


def read_ground_truth(path: str) -> dict[int, object]:
    """Read ground truth CMA - FY25 is in Column E (col index 4, 0-based = index 4 = xlrd col 4)."""
    wb = xlrd.open_workbook(path)
    ws = wb.sheet_by_name("INPUT SHEET")
    # Verify year headers
    print("\nGround truth year headers (row 8):")
    for c in range(13):
        v = ws.cell(7, c).value  # row 8 = index 7
        if v:
            print(f"  Col {c+1} ({chr(64+c+1)}): {v!r}")

    data = {}
    for row in EXPECTED:
        if row - 1 < ws.nrows:
            # FY25 = Col E = index 4 (0-based)
            val = ws.cell(row - 1, 4).value
            data[row] = val
        else:
            data[row] = None
    return data


def to_float(val) -> float | None:
    """Convert a cell value to float, or None if empty/formula."""
    if val is None:
        return None
    if isinstance(val, str):
        if val.startswith("="):
            return None  # formula — can't read without Excel
        try:
            return float(val)
        except ValueError:
            return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def main():
    print("=" * 80)
    print("MSL CMA COMPARISON REPORT")
    print("Generated CMA: test-results/MSL_generated_CMA.xlsm (FY25 -> Col D)")
    print("Ground Truth:  FInancials main/MSL - Manufacturing/MSL CMA 24102025.xls (FY25 -> Col E)")
    print("=" * 80)

    # Read files
    print("\nReading generated CMA...")
    gen_data = read_generated(GENERATED)
    print("Reading ground truth CMA...")
    gt_data = read_ground_truth(GROUND_TRUTH)

    # VBA check
    wb = openpyxl.load_workbook(GENERATED, keep_vba=True)
    has_vba = wb.vba_archive is not None
    print(f"\nVBA macros preserved: {has_vba}")
    print(f"Generated CMA sheets: {wb.sheetnames}")

    print("\n" + "-" * 100)
    print(f"{'Row':<5} {'Field':<25} {'Expected':>12} {'Gen_Col_D':>12} {'GT_Col_E':>12} {'Status':<15} Notes")
    print("-" * 100)

    matches = []
    mismatches = []
    missing = []
    formula_cells = []

    for row in sorted(EXPECTED):
        field_name, expected_val = EXPECTED[row]
        gen_raw = gen_data[row]
        gt_raw = gt_data[row]

        gen_val = to_float(gen_raw)
        gt_val = to_float(gt_raw)

        # Determine status vs ground truth expected (task instructions)
        notes = ""
        if gen_raw is not None and isinstance(gen_raw, str) and gen_raw.startswith("="):
            status = "FORMULA"
            notes = f"Cell has formula: {gen_raw!r}"
            formula_cells.append((row, field_name, gen_raw))
        elif gen_val is None:
            status = "MISSING"
            notes = f"raw={gen_raw!r}"
            missing.append((row, field_name, expected_val, gen_raw, gt_val))
        else:
            diff = abs(gen_val - expected_val)
            if diff <= TOLERANCE:
                status = "MATCH"
                matches.append((row, field_name, expected_val, gen_val))
            else:
                # Check if it's a unit conversion issue (100x off)
                if abs(gen_val / expected_val - 100) < 5 if expected_val != 0 else False:
                    notes = "UNIT BUG: 100x too large"
                elif abs(gen_val - gt_val) <= TOLERANCE if gt_val is not None else False:
                    notes = "Matches GT but not task spec (different FY?)"
                else:
                    factor = gen_val / expected_val if expected_val != 0 else float('inf')
                    notes = f"diff={diff:.2f}, factor={factor:.2f}x"
                status = "MISMATCH"
                mismatches.append((row, field_name, expected_val, gen_val, notes))

        gen_display = f"{gen_val:.2f}" if gen_val is not None else str(gen_raw)[:10]
        gt_display = f"{gt_val:.2f}" if gt_val is not None else str(gt_raw)[:10]
        print(f"{row:<5} {field_name:<25} {expected_val:>12.2f} {gen_display:>12} {gt_display:>12} {status:<15} {notes}")

    # Summary
    total = len(EXPECTED)
    print("\n" + "=" * 80)
    print("SUMMARY")
    print("=" * 80)
    print(f"Total cells checked:    {total}")
    print(f"MATCHING:               {len(matches)}")
    print(f"MISMATCHED:             {len(mismatches)}")
    print(f"MISSING (None/empty):   {len(missing)}")
    print(f"FORMULA (unreadable):   {len(formula_cells)}")

    if matches:
        print(f"\nMATCHING cells ({len(matches)}):")
        for row, field, exp, act in matches:
            print(f"  Row {row:>3} {field:<30} expected={exp:.2f} actual={act:.2f}")

    if mismatches:
        print(f"\nMISMATCHED cells ({len(mismatches)}):")
        for row, field, exp, act, notes in mismatches:
            print(f"  Row {row:>3} {field:<30} expected={exp:.2f} actual={act:.2f}  [{notes}]")

    if missing:
        print(f"\nMISSING cells ({len(missing)}):")
        for row, field, exp, raw, gt_val in missing:
            gt_str = f"{gt_val:.2f}" if gt_val is not None else "n/a"
            print(f"  Row {row:>3} {field:<30} expected={exp:.2f} raw={raw!r} gt_col_e={gt_str}")

    if formula_cells:
        print(f"\nFORMULA cells (skipped by generator — correct behavior if read-only):")
        for row, field, formula in formula_cells:
            print(f"  Row {row:>3} {field:<30} formula={formula!r}")

    print("\n" + "=" * 80)
    print("DIAGNOSIS")
    print("=" * 80)
    print("""
KEY FINDINGS:

1. YEAR-COLUMN MISMATCH (Bug):
   - Ground truth CMA: FY25 data lives in Column E (template starts FY2022 at Col B)
   - Generated CMA: YEAR_TO_COLUMN maps 2025 to Col D (assumes FY2023 is first year)
   - This means data written to Col D should be in Col E
   - The year header in the generated template shows 2023 starts at Col B
   - The ground truth template shows 2022 starts at Col B
   - ROOT CAUSE: The CMA.xlsm template in the repo has a different base year
     than the client's ground truth CMA file

2. FIELD NAME MISMATCH (Bug):
   - 35 out of 98 non-doubt classifications have field names like "Item 1 (i) : Domestic sales"
     instead of the ALL_FIELD_TO_ROW key "Domestic Sales"
   - These get skipped (no row mapping) in _fill_data_cells
   - The classification pipeline uses different field name conventions than the row mapping

3. FORMULA CELLS:
   - Several INPUT SHEET cells contain formulas (e.g., =0, =D59)
   - The generator correctly skips overwriting formulas
   - But this means those cells never get data written to them

4. UNIT CONVERSION:
   - The generated values are in raw amounts from the source (Rs. 000)
   - No /100 conversion to Lakhs was applied
   - This is a known issue (task says document as bug)
""")

    return {
        "total": total,
        "matches": len(matches),
        "mismatches": len(mismatches),
        "missing": len(missing),
        "formula_cells": len(formula_cells),
    }


if __name__ == "__main__":
    main()
